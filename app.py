import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
from datetime import datetime
import io
import os
from openai import OpenAI
from anthropic import Anthropic
from google import genai as google_genai
from fpdf import FPDF
from openpyxl.styles import Font

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.set_page_config(
    page_title="Full Turn Analytics | Portfolio Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Authentication gate ──
def check_password():
    """Return True if the user has entered the correct password."""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    pwd = st.text_input("Password", type="password", placeholder="Enter dashboard password")
    if pwd:
        if pwd == st.secrets.get("password", ""):
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not check_password():
    st.stop()

YEARS = [2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
TREND_YEARS = [2021, 2022, 2023, 2024, 2025]
CHART_TEMPLATE = "plotly_white"

# ── Materials vs Labor category classification ──
MATERIALS_CATS = [
    "Supplies", "Appliances", "Flooring Materials",
    "Cabinets Materials", "Countertops Materials", "Windows",
]
LABOR_CATS = [
    "Labor General", "Flooring Labor", "Electric General",
    "Countertops Labor", "Plumbing", "Powerwash and Demo",
    "Management Fee", "Scrape Ceiling", "Glaze", "Cabinets Labor", "Paint",
]

# ── Expense analysis constants ──
EXPENSE_YEARS = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
CORE_LABOR = [
    "Flooring Labor", "Countertops Labor", "Cabinets Labor",
    "Paint", "Glaze", "Labor General", "Management Fee",
]
CORE_MATERIALS = [
    "Flooring Materials", "Countertops Materials",
    "Cabinets Materials", "Appliances",
]
OTHER_CATS = [
    "Electric General", "Plumbing", "Scrape Ceiling",
    "Supplies", "Windows", "Powerwash and Demo",
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STYLING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .block-container { padding-top: 1.5rem; }

    /* Sidebar */
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%); }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h3 { color: #ffffff !important; }

    /* KPI cards */
    div[data-testid="stMetric"] {
        background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px;
        padding: 18px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        border-left: 4px solid #2563eb;
    }
    div[data-testid="stMetric"] label {
        font-size: 11px !important; text-transform: uppercase;
        letter-spacing: 0.6px; color: #64748b !important; font-weight: 600 !important;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        font-size: 26px !important; font-weight: 700 !important; color: #0f172a !important;
    }

    /* Tables */
    .stDataFrame th {
        background: #f1f5f9 !important; font-weight: 600 !important;
        text-transform: uppercase; font-size: 11px !important;
        letter-spacing: 0.5px; color: #475569 !important;
    }
    .stDataFrame td { font-size: 13px !important; }

    /* Headers */
    .page-banner {
        background: linear-gradient(135deg, #0f172a, #1e40af);
        color: white; padding: 28px 32px; border-radius: 12px; margin-bottom: 24px;
    }
    .page-banner h1 { color: white !important; margin: 0 !important; font-size: 26px !important; font-weight: 700; }
    .page-banner p  { color: #93c5fd; margin: 4px 0 0 0; font-size: 13px; }

    .section-bar {
        border-left: 4px solid #2563eb; background: #f8fafc;
        padding: 10px 16px; margin: 28px 0 16px 0; border-radius: 0 8px 8px 0;
    }
    .section-bar h3 { margin: 0 !important; font-size: 15px !important; color: #1e293b !important; font-weight: 600; }

    .insight-box {
        background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 10px;
        padding: 16px 20px; margin: 8px 0 16px 0; font-size: 13.5px; line-height: 1.7; color: #1e293b;
    }
    .insight-box strong { color: #1e40af; }

    .risk-flag {
        border-radius: 8px; padding: 12px 16px; margin: 6px 0; font-size: 13px;
    }
    .risk-flag-high {
        background: #fef2f2; border: 1px solid #fecaca; border-left: 4px solid #dc2626;
    }
    .risk-flag-medium {
        background: #fffbeb; border: 1px solid #fde68a; border-left: 4px solid #f59e0b;
    }

    /* Tabs */
    .stTabs [data-baseweb="tab"] { font-weight: 600; font-size: 14px; }
    .stTabs [aria-selected="true"] { border-bottom-color: #2563eb !important; }

    /* Footer */
    .dashboard-footer {
        margin-top: 60px; padding: 20px 0; border-top: 1px solid #e2e8f0;
        text-align: center; font-size: 11px; color: #94a3b8; letter-spacing: 0.3px;
    }

    /* Sidebar active view accent */
    [data-testid="stSidebar"] [role="radiogroup"] label[data-checked="true"] {
        background: rgba(37,99,235,0.15) !important; border-radius: 6px;
    }

    /* ── Scope Checklist tables ── */
    .scope-checklist { width: 100%; border-collapse: collapse; margin: 0 0 8px 0; font-family: 'Inter', sans-serif; }
    .scope-checklist th {
        text-align: left; font-size: 11px; text-transform: uppercase;
        letter-spacing: 0.5px; color: #64748b; font-weight: 600;
        padding: 8px 12px; border-bottom: 2px solid #e2e8f0; background: #f8fafc;
    }
    .scope-checklist th:last-child { text-align: right; }
    .scope-checklist .group-hdr td {
        font-weight: 700; font-size: 12px; text-transform: uppercase;
        letter-spacing: 0.5px; padding: 10px 12px 6px 12px; border-bottom: none;
    }
    .scope-checklist .group-hdr-labor td   { color: #1e40af; border-top: 3px solid #2563eb; }
    .scope-checklist .group-hdr-matl td    { color: #047857; border-top: 3px solid #10b981; }
    .scope-checklist .group-hdr-other td   { color: #6d28d9; border-top: 3px solid #7c3aed; }
    .scope-checklist .row-done td {
        padding: 6px 12px; font-size: 13px; color: #1e293b;
        border-bottom: 1px solid #f1f5f9;
    }
    .scope-checklist .row-done .status { color: #16a34a; font-weight: 700; }
    .scope-checklist .row-done td:last-child { text-align: right; font-weight: 600; font-variant-numeric: tabular-nums; }
    .scope-checklist .row-skip td {
        padding: 6px 12px; font-size: 13px; color: #94a3b8;
        border-bottom: 1px solid #f1f5f9;
    }
    .scope-checklist .row-skip .status { color: #cbd5e1; }
    .scope-checklist .row-skip td:last-child { text-align: right; }

    /* Section wrappers for Pre-Scoping vs Projected */
    .scope-card {
        border-radius: 10px; padding: 20px 24px; margin: 12px 0 20px 0;
    }
    .scope-card-history {
        background: #f0f7ff; border: 1px solid #bfdbfe; border-left: 5px solid #2563eb;
    }
    .scope-card-projected {
        background: #fffbeb; border: 1px solid #fde68a; border-left: 5px solid #f59e0b;
    }
    .scope-card .scope-title {
        font-size: 14px; font-weight: 700; margin: 0 0 4px 0;
    }
    .scope-card-history .scope-title { color: #1e40af; }
    .scope-card-projected .scope-title { color: #92400e; }
    .scope-card .scope-subtitle {
        font-size: 12px; margin: 0 0 12px 0; color: #64748b;
    }

    /* Multi-year history table */
    .scope-checklist th.year-col {
        text-align: right; min-width: 90px; font-variant-numeric: tabular-nums;
    }
    .scope-checklist .total-row td {
        font-weight: 700; font-size: 13px; border-bottom: 2px solid #cbd5e1;
        background: #f1f5f9; padding: 8px 12px;
    }
    .scope-checklist .total-row td:first-child { text-align: left; }
    .scope-checklist td.year-val {
        text-align: right; font-variant-numeric: tabular-nums; padding: 6px 12px;
    }
    .scope-checklist .row-done td.year-val { font-weight: 600; }
    .scope-checklist .row-skip td.year-val { color: #cbd5e1; }
    .scope-checklist .year-subtext {
        display: block; font-size: 9px; color: #94a3b8; font-weight: 400;
        text-transform: none; letter-spacing: 0;
    }

    /* Projected scope — recently-done row state */
    .scope-checklist .row-recent td {
        padding: 6px 12px; font-size: 13px; color: #92400e;
        border-bottom: 1px solid #f1f5f9; background: #fffbeb;
    }
    .scope-checklist .row-recent .status { color: #f59e0b; font-weight: 700; }
    .scope-checklist .row-recent td.year-val {
        text-align: right; text-decoration: line-through; color: #94a3b8;
    }

</style>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def banner(title, subtitle):
    st.markdown(f'<div class="page-banner"><h1>{title}</h1><p>{subtitle}</p></div>', unsafe_allow_html=True)

def section(title):
    st.markdown(f'<div class="section-bar"><h3>{title}</h3></div>', unsafe_allow_html=True)

def insight(text):
    st.markdown(f'<div class="insight-box">{text}</div>', unsafe_allow_html=True)

def fmt(val, decimals=0):
    if pd.isna(val):
        return "—"
    if val < 0:
        return f"-${abs(val):,.{decimals}f}"
    return f"${val:,.{decimals}f}"

def pct(val):
    if pd.isna(val):
        return "—"
    return f"{val:+.1f}%"

def expense_year_label(yr):
    return "2026 YTD" if yr == 2026 else str(yr)

def _pdf(s):
    """Sanitize text for FPDF's default latin-1 font (replace em/en dashes)."""
    return str(s).replace("\u2014", "-").replace("\u2013", "-")

EXPENSE_YEAR_LABELS = [expense_year_label(y) for y in EXPENSE_YEARS]

def footer():
    _data_date = ft_turns["Move-Out Date"].max()
    _date_label = _data_date.strftime("%b %Y") if pd.notna(_data_date) else "N/A"
    st.markdown(
        '<div class="dashboard-footer">'
        'CONFIDENTIAL — Full Turn Analytics Dashboard &nbsp;|&nbsp; '
        f'Data through {_date_label} &nbsp;|&nbsp; '
        'Prepared for Executive Review'
        '</div>',
        unsafe_allow_html=True,
    )


def render_scope_history_table(unit_df, unit_ts):
    """Build an HTML table showing all historical work on a unit by year and category.

    Rows = 17 budget categories in standard group order.
    Columns = years with actual turn data, most recent first.
    """
    ALL_CATS = CORE_LABOR + CORE_MATERIALS + OTHER_CATS
    GROUP_CSS = {"Core Labor": "group-hdr-labor", "Core Materials": "group-hdr-matl", "Other": "group-hdr-other"}

    work = unit_df.copy()
    work["Year"] = work["Move-Out Date"].dt.year

    # Pivot: category × year
    pivot = work.pivot_table(
        index="Budget Category", columns="Year",
        values="Invoice Amount", aggfunc="sum", fill_value=0,
    )

    # Only keep years with non-zero total spend
    year_cols = sorted([y for y in pivot.columns if pivot[y].sum() != 0], reverse=True)
    if not year_cols:
        return "<p style='color:#64748b;'>No turn history available.</p>", [], {}

    # Turn type labels per year
    ts_copy = unit_ts.copy()
    ts_copy["Year"] = ts_copy["Move-Out Date"].dt.year
    type_labels = ts_copy.groupby("Year")["Turn Type"].apply(lambda x: " / ".join(sorted(x.unique()))).to_dict()

    year_totals = {y: pivot[y].sum() for y in year_cols}

    # Build header row
    year_hdrs = "".join(
        f'<th class="year-col">{int(y)}<span class="year-subtext">{type_labels.get(y, "")}</span></th>'
        for y in year_cols
    )
    header = f'<tr><th style="width:32px;"></th><th>Category</th>{year_hdrs}</tr>'

    # Total row
    total_cells = "".join(f'<td class="year-val">{fmt(year_totals[y])}</td>' for y in year_cols)
    total_row = f'<tr class="total-row"><td></td><td>Total Per Unit</td>{total_cells}</tr>'

    # Category rows
    rows_html = [total_row]
    for group_label, group_cats in [("Core Labor", CORE_LABOR), ("Core Materials", CORE_MATERIALS), ("Other", OTHER_CATS)]:
        css = GROUP_CSS[group_label]
        colspan = 2 + len(year_cols)
        rows_html.append(f'<tr class="group-hdr {css}"><td colspan="{colspan}">{group_label}</td></tr>')
        for c in group_cats:
            has_any = any(pivot.loc[c, y] > 0 for y in year_cols) if c in pivot.index else False
            row_class = "row-done" if has_any else "row-skip"
            status = "&#10003;" if has_any else "—"
            cells = ""
            for y in year_cols:
                val = pivot.loc[c, y] if c in pivot.index and y in pivot.columns else 0
                if val > 0:
                    cells += f'<td class="year-val">{fmt(val)}</td>'
                elif val < 0:
                    cells += f'<td class="year-val">{fmt(val)}</td>'
                else:
                    cells += f'<td class="year-val" style="color:#cbd5e1;">—</td>'
            rows_html.append(
                f'<tr class="{row_class}">'
                f'<td class="status">{status}</td>'
                f'<td>{c}</td>'
                f'{cells}'
                f'</tr>'
            )

    html = (
        f'<table class="scope-checklist">'
        f'<thead>{header}</thead>'
        f'<tbody>{"".join(rows_html)}</tbody></table>'
    )
    return html, year_cols, year_totals


def render_projected_scope_table(proj_amounts, excluded_recent, last_done_year,
                                 amount_label="Projected Cost",
                                 comp_lines=None, comp_info=None):
    """Build an HTML checklist for projected scope with recently-done flagging
    and optional comp turn columns.

    Three row states:
      row-done (green ✓): Category projected
      row-skip (gray —): Not projected
      row-recent (amber ↻): Projected by comps but recently done on this unit

    comp_lines : DataFrame of line items for comp turns (optional)
    comp_info  : DataFrame with Turn Key, Unit Label, Move-Out Date, Turn Type (optional)
    """
    GROUP_CSS = {"Core Labor": "group-hdr-labor", "Core Materials": "group-hdr-matl", "Other": "group-hdr-other"}
    ALL_CATS = CORE_LABOR + CORE_MATERIALS + OTHER_CATS

    # Build comp column data if provided
    has_comps = comp_lines is not None and comp_info is not None and len(comp_info) > 0
    turn_keys = []
    comp_pivot = pd.DataFrame()
    comp_totals = {}
    if has_comps:
        comp_info = comp_info.sort_values("Move-Out Date", ascending=False).reset_index(drop=True)
        turn_keys = comp_info["Turn Key"].tolist()
        comp_pivot = comp_lines.pivot_table(
            index="Budget Category", columns="Turn Key",
            values="Invoice Amount", aggfunc="sum", fill_value=0,
        )
        comp_totals = {tk: comp_pivot[tk].sum() if tk in comp_pivot.columns else 0 for tk in turn_keys}

    n_extra = len(turn_keys)
    total_cols = 3 + n_extra  # status + category + projected + comp columns

    # Build comp column headers
    comp_hdrs = ""
    for _, row in (comp_info.iterrows() if has_comps else pd.DataFrame().iterrows()):
        tk = row["Turn Key"]
        unit_lbl = row["Unit Label"]
        date_str = row["Move-Out Date"].strftime("%b %y") if pd.notna(row["Move-Out Date"]) else "—"
        turn_type_short = row.get("Turn Type", "")
        if turn_type_short == "Make Ready":
            tt_abbr = "MR"
        elif turn_type_short == "Full Turn":
            tt_abbr = "FT"
        elif turn_type_short == "Partial Turn":
            tt_abbr = "PT"
        else:
            tt_abbr = turn_type_short
        comp_hdrs += (
            f'<th class="year-col">{unit_lbl}'
            f'<span class="year-subtext">{tt_abbr} · {date_str}</span></th>'
        )

    proj_col_style = 'style="background-color:#f1f5f9;"'
    header = (
        f'<tr><th style="width:32px;"></th><th>Category</th>'
        f'<th {proj_col_style}>{amount_label}</th>{comp_hdrs}</tr>'
    )

    # Projected total
    proj_total = sum(proj_amounts.get(c, 0) for c in ALL_CATS if c not in excluded_recent)

    # Total row
    total_cells = f'<td class="year-val" {proj_col_style}>{fmt(proj_total)}</td>'
    for tk in turn_keys:
        total_cells += f'<td class="year-val">{fmt(comp_totals.get(tk, 0))}</td>'
    total_row = f'<tr class="total-row"><td></td><td>Total</td>{total_cells}</tr>'

    # Category rows
    rows_html = [total_row]
    for group_label, group_cats in [("Core Labor", CORE_LABOR), ("Core Materials", CORE_MATERIALS), ("Other", OTHER_CATS)]:
        css = GROUP_CSS[group_label]
        rows_html.append(f'<tr class="group-hdr {css}"><td colspan="{total_cols}">{group_label}</td></tr>')
        for c in group_cats:
            amt = proj_amounts.get(c, 0) if hasattr(proj_amounts, "get") else 0
            # Build comp cells
            comp_cells = ""
            for tk in turn_keys:
                val = comp_pivot.loc[c, tk] if c in comp_pivot.index and tk in comp_pivot.columns else 0
                if val > 0:
                    comp_cells += f'<td class="year-val">{fmt(val)}</td>'
                else:
                    comp_cells += f'<td class="year-val" style="color:#cbd5e1;">—</td>'

            if c in excluded_recent:
                comp_amt = excluded_recent[c]
                yr = last_done_year.get(c, "")
                rows_html.append(
                    f'<tr class="row-recent">'
                    f'<td class="status">&#8635;</td>'
                    f'<td>{c} <span style="font-size:11px;color:#94a3b8;">— done in {yr}</span></td>'
                    f'<td class="year-val" {proj_col_style}>{fmt(comp_amt)}</td>'
                    f'{comp_cells}'
                    f'</tr>'
                )
            elif amt > 0:
                rows_html.append(
                    f'<tr class="row-done">'
                    f'<td class="status">&#10003;</td>'
                    f'<td>{c}</td>'
                    f'<td class="year-val" {proj_col_style}>{fmt(amt)}</td>'
                    f'{comp_cells}'
                    f'</tr>'
                )
            else:
                rows_html.append(
                    f'<tr class="row-skip">'
                    f'<td class="status">—</td>'
                    f'<td>{c}</td>'
                    f'<td {proj_col_style}>—</td>'
                    f'{comp_cells}'
                    f'</tr>'
                )
    return (
        f'<table class="scope-checklist">'
        f'<thead>{header}</thead>'
        f'<tbody>{"".join(rows_html)}</tbody></table>'
    )


def render_floor_plan_comparison_table(fp_lines, fp_turns_info):
    """Build an HTML table comparing category costs across comparable same-floor-plan turns.

    fp_lines: DataFrame of invoice line items for the comparable turns
    fp_turns_info: DataFrame with Turn Key, Unit Label, Move-Out Date, total_cost (one row per turn)
    Returns HTML string.
    """
    ALL_CATS = CORE_LABOR + CORE_MATERIALS + OTHER_CATS
    GROUP_CSS = {"Core Labor": "group-hdr-labor", "Core Materials": "group-hdr-matl", "Other": "group-hdr-other"}

    if len(fp_turns_info) == 0:
        return "<p style='color:#64748b;'>No comparable floor plan turns found.</p>"

    # Sort turns by date descending
    fp_turns_info = fp_turns_info.sort_values("Move-Out Date", ascending=False).reset_index(drop=True)
    turn_keys = fp_turns_info["Turn Key"].tolist()

    # Pivot: category × Turn Key
    pivot = fp_lines.pivot_table(
        index="Budget Category", columns="Turn Key",
        values="Invoice Amount", aggfunc="sum", fill_value=0,
    )

    turn_totals = {tk: pivot[tk].sum() if tk in pivot.columns else 0 for tk in turn_keys}

    # Build header — "Unit XXX\nMon YY"
    col_hdrs = ""
    for _, row in fp_turns_info.iterrows():
        tk = row["Turn Key"]
        unit_lbl = row["Unit Label"]
        date_str = row["Move-Out Date"].strftime("%b %y") if pd.notna(row["Move-Out Date"]) else "—"
        turn_type_short = row.get("Turn Type", "")
        col_hdrs += (
            f'<th class="year-col">Unit {unit_lbl}'
            f'<span class="year-subtext">{turn_type_short} · {date_str}</span></th>'
        )
    header = f'<tr><th style="width:32px;"></th><th>Category</th>{col_hdrs}</tr>'

    # Total row
    total_cells = "".join(f'<td class="year-val">{fmt(turn_totals[tk])}</td>' for tk in turn_keys)
    total_row = f'<tr class="total-row"><td></td><td>Total Per Unit</td>{total_cells}</tr>'

    # Category rows
    rows_html = [total_row]
    for group_label, group_cats in [("Core Labor", CORE_LABOR), ("Core Materials", CORE_MATERIALS), ("Other", OTHER_CATS)]:
        css = GROUP_CSS[group_label]
        colspan = 2 + len(turn_keys)
        rows_html.append(f'<tr class="group-hdr {css}"><td colspan="{colspan}">{group_label}</td></tr>')
        for c in group_cats:
            has_any = any((pivot.loc[c, tk] > 0 if c in pivot.index and tk in pivot.columns else False) for tk in turn_keys)
            row_class = "row-done" if has_any else "row-skip"
            status = "&#10003;" if has_any else "—"
            cells = ""
            for tk in turn_keys:
                val = pivot.loc[c, tk] if c in pivot.index and tk in pivot.columns else 0
                if val > 0:
                    cells += f'<td class="year-val">{fmt(val)}</td>'
                else:
                    cells += f'<td class="year-val" style="color:#cbd5e1;">—</td>'
            rows_html.append(
                f'<tr class="{row_class}">'
                f'<td class="status">{status}</td>'
                f'<td>{c}</td>'
                f'{cells}'
                f'</tr>'
            )

    return (
        f'<table class="scope-checklist">'
        f'<thead>{header}</thead>'
        f'<tbody>{"".join(rows_html)}</tbody></table>'
    )


def render_category_table(title, categories, data, years=None, year_labels=None, yoy_pairs=None):
    """Render a pivot table for a set of budget categories with YoY change column(s).

    The Total row is included in the DataFrame and styled with a subtle
    background via Pandas Styler.

    Parameters
    ----------
    title : str – section title
    categories : list – budget category names to include
    data : DataFrame – must have columns: Budget Category, Year, avg_per_turn
    years : list[int] – year columns (default EXPENSE_YEARS)
    year_labels : list[str] – display labels for year columns (default EXPENSE_YEAR_LABELS)
    yoy_pairs : list[tuple] – explicit (prior_year, current_year) pairs for YoY columns.
                               If None, auto-generates a single column from the last two years.
    """
    if years is None:
        years = EXPENSE_YEARS
    if year_labels is None:
        year_labels = EXPENSE_YEAR_LABELS

    st.markdown(f"**{title}**")
    subset = data[data["Budget Category"].isin(categories)]
    if len(subset) == 0:
        st.info(f"No data for {title.split(' (')[0]} categories.")
        return
    pivot = subset.pivot_table(
        index="Budget Category", columns="Year",
        values="avg_per_turn", fill_value=0
    ).reindex(columns=years, fill_value=0)
    pivot = pivot.reindex([c for c in categories if c in pivot.index])
    pivot = pivot.fillna(0)
    if len(pivot) == 0:
        st.info(f"No data for {title.split(' (')[0]} categories.")
        return
    pivot = pivot.loc[pivot.sum(axis=1) > 0]
    if len(pivot) == 0:
        st.info(f"No data for {title.split(' (')[0]} categories.")
        return

    # Add Total row
    pivot.loc["Total"] = pivot.sum()

    # Build YoY pairs list
    if yoy_pairs is None and len(years) >= 2:
        yoy_pairs = [(years[-2], years[-1])]
    elif yoy_pairs is None:
        yoy_pairs = []

    # Compute YoY columns
    yoy_columns = {}
    for prior, current in yoy_pairs:
        if prior in pivot.columns and current in pivot.columns:
            yoy_vals = pivot[current] / pivot[prior].replace(0, np.nan) - 1
            label = f"'{str(prior)[-2:]}→'{str(current)[-2:]}"
            yoy_columns[label] = yoy_vals

    # Format for display
    pivot_display = pivot.copy()
    pivot_display.columns = year_labels
    for col in pivot_display.columns:
        pivot_display[col] = pivot_display[col].apply(fmt)
    for label, yoy_vals in yoy_columns.items():
        pivot_display[label] = yoy_vals.apply(
            lambda x: f"{x:+.0%}" if pd.notna(x) and np.isfinite(x) else "—"
        )

    # Style: Total row + YoY column tint
    yoy_col_labels = list(yoy_columns.keys())

    def apply_styles(styler):
        styler.set_properties(
            subset=pd.IndexSlice["Total", :],
            **{"background-color": "#f1f5f9", "font-weight": "700", "border-top": "2px solid #cbd5e1"},
        )
        if yoy_col_labels:
            styler.set_properties(
                subset=pd.IndexSlice[:, yoy_col_labels],
                **{"background-color": "#f8fafc"},
            )
            styler.set_properties(
                subset=pd.IndexSlice[:, yoy_col_labels[0]],
                **{"border-left": "2px solid #e2e8f0"},
            )
        return styler

    st.dataframe(apply_styles(pivot_display.style), use_container_width=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA LOADING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@st.cache_data
def load_data():
    path = Path(__file__).parent / "Unit Turns - AI Clean - 2.26.2026.xlsx"
    df = pd.read_excel(path)

    df["Building Code"] = df["Building Code"].fillna("").astype(str).str.strip()
    df["Unit Number"]   = df["Unit Number"].fillna("").astype(str).str.strip()
    df["Property ID"]   = df["Property ID"].fillna("").astype(str).str.strip()
    df["Invoice Amount"] = pd.to_numeric(df["Invoice Amount"], errors="coerce").fillna(0)
    df["Move-Out Date"]  = pd.to_datetime(df["Move-Out Date"], errors="coerce")
    df["Invoice Date"]   = pd.to_datetime(df["Invoice Date"], errors="coerce")

    # Cap obviously wrong dates
    cap = pd.Timestamp(datetime.now().year + 1, 12, 31)
    df.loc[df["Invoice Date"] > cap, "Invoice Date"] = pd.NaT

    df = df[df["Property ID"] != ""].copy()

    # ── Unique Unit Key ──
    def uid(r):
        if r["Building Code"] != "":
            return f"{r['Property ID']}|{r['Building Code']}|{r['Unit Number']}"
        return f"{r['Property ID']}|{r['Unit Number']}"

    def uid_display(r):
        if r["Building Code"] != "":
            return f"{r['Building Code']} | {r['Unit Number']}"
        return r["Unit Number"]

    df["UID"]         = df.apply(uid, axis=1)
    df["Unit Label"]  = df.apply(uid_display, axis=1)
    df["Turn Key"]    = df["UID"] + "|" + df["Move-Out Date"].astype(str)

    # ── Full Turn only ──
    ft = df[df["Turn Type"] == "Full Turn"].copy()
    ft["Year"] = ft["Move-Out Date"].dt.year

    # ── Turn-level summary ──
    turns = (
        ft.groupby(["Property ID", "Property Name", "UID", "Unit Label",
                     "Move-Out Date", "Turn Key", "Floor Plan",
                     "Bedrooms", "Bathrooms", "Year"])
        .agg(total_cost=("Invoice Amount", "sum"),
             line_items=("Invoice Amount", "count"),
             completion_date=("Invoice Date", "max"))
        .reset_index()
    )
    turns["Duration"] = (turns["completion_date"] - turns["Move-Out Date"]).dt.days
    turns.loc[turns["Duration"] < 0, "Duration"] = pd.NA

    return df, ft, turns


_df_all, ft_lines, ft_turns = load_data()
PROPERTY_ORDER = [
    "Monterey Park", "Woodman", "Collins", "Lindley", "El Rancho",
    "51 at the Village", "Alta Vista", "Roscoe", "Topanga", "Darby",
    "Fruitland", "Dickens", "Garfield", "Woodbridge",
    "12756 Moorpark", "12800 Moorpark",
]
_all_props = set(ft_turns["Property Name"].unique())
PROPERTIES = [p for p in PROPERTY_ORDER if p in _all_props] + sorted(_all_props - set(PROPERTY_ORDER))
_PROP_RANK = {name: i for i, name in enumerate(PROPERTIES)}

def prop_sort_key(names):
    """Return sort keys for a Series/Index of property names using PROPERTY_ORDER."""
    return names.map(lambda n: _PROP_RANK.get(n, 999))


@st.cache_data
def build_turn_summary(source_df):
    return (
        source_df.groupby(["Property ID", "Property Name", "UID",
                           "Unit Label", "Move-Out Date", "Turn Type",
                           "Floor Plan", "Bedrooms", "Bathrooms", "Turn Key"])
        .agg(total_cost=("Invoice Amount", "sum"),
             line_items=("Invoice Amount", "count"),
             earliest_invoice=("Invoice Date", "min"),
             latest_invoice=("Invoice Date", "max"))
        .reset_index()
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SIDEBAR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.sidebar.markdown("### FULL TURN ANALYTICS")
st.sidebar.caption("Portfolio Renovation Intelligence Platform")
st.sidebar.markdown("---")

view = st.sidebar.radio("View", [
    "1 — Executive Summary",
    "2 — Portfolio Overview",
    "3 — Property Summary",
    "4 — Unit Search",
    "5 — Rent Roll",
    "6 — Data Health",
    "7 — AI Data Review",
])

st.sidebar.markdown("---")
st.sidebar.caption(
    f"**{len(ft_turns):,}** Full Turns  •  **{ft_turns['Property ID'].nunique()}** Properties  •  "
    f"**{ft_turns['UID'].nunique()}** Units  •  **{fmt(ft_turns['total_cost'].sum())}** Total Spend"
)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 3: PROPERTY SUMMARY
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if view == "3 — Property Summary":
    banner("Property Summary", "Single-property deep dive — turn volume, expense breakdown, and category analysis")

    prop = st.selectbox("Select Property", PROPERTIES)
    p_turns = ft_turns[ft_turns["Property Name"] == prop].copy()
    p_lines = ft_lines[ft_lines["Property Name"] == prop].copy()

    # KPIs
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Full Turns", f"{len(p_turns):,}")
    c2.metric("Total Spend", fmt(p_turns["total_cost"].sum()))
    c3.metric("Avg Cost", fmt(p_turns["total_cost"].mean()))
    c4.metric("Median Cost", fmt(p_turns["total_cost"].median()))
    dur = p_turns["Duration"].dropna()
    c5.metric("Median Duration", f"{dur.median():.0f} days" if len(dur) else "—")

    # ── YoY Count ──
    section("Year-over-Year Full Turn Count (2016 – 2025)")

    yoy = p_turns[p_turns["Year"].isin(YEARS)].groupby("Year").size().reindex(YEARS, fill_value=0)
    yoy_df = pd.DataFrame({"Year": YEARS, "Full Turns": yoy.values})

    col1, col2 = st.columns([2, 3])
    with col1:
        # Add YoY delta
        yoy_df["vs Prior Year"] = yoy_df["Full Turns"].diff()
        yoy_df["vs Prior Year"] = yoy_df["vs Prior Year"].apply(
            lambda x: f"+{x:.0f}" if pd.notna(x) and x > 0 else (f"{x:.0f}" if pd.notna(x) and x != 0 else "—")
        )
        yoy_df["Year"] = yoy_df["Year"].astype(str)
        st.dataframe(yoy_df, use_container_width=True, hide_index=True)

    with col2:
        fig = px.bar(
            yoy_df, x="Year", y="Full Turns", text="Full Turns",
            template=CHART_TEMPLATE, color_discrete_sequence=["#2563eb"],
        )
        fig.update_traces(textposition="outside", textfont_size=13)
        fig.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=280,
                          xaxis_title="", yaxis_title="Turns")
        st.plotly_chart(fig, use_container_width=True)

    # ── Turns by Floor Plan ──
    section("Full Turns by Floor Plan")

    fp_data = p_turns.groupby("Floor Plan").agg(
        turns=("Turn Key", "count"),
        avg_cost=("total_cost", "mean"),
        total=("total_cost", "sum"),
    ).reset_index().sort_values("Floor Plan")

    col1, col2 = st.columns([2, 3])
    with col1:
        fp_display = fp_data.copy()
        fp_display["avg_cost"] = fp_display["avg_cost"].apply(fmt)
        fp_display["total"] = fp_display["total"].apply(fmt)
        fp_display.columns = ["Floor Plan", "Turns", "Avg Cost", "Total Spend"]
        st.dataframe(fp_display, use_container_width=True, hide_index=True)

    with col2:
        fig_fp = px.bar(
            fp_data, x="Floor Plan", y="turns", text="turns",
            template=CHART_TEMPLATE, color_discrete_sequence=["#0ea5e9"],
        )
        fig_fp.update_traces(textposition="outside")
        fig_fp.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=280,
                             xaxis_title="", yaxis_title="Turns")
        st.plotly_chart(fig_fp, use_container_width=True)

    # ══════════════════════════════════════════════════
    # EXPENSE SUMMARY BY CATEGORY GROUP
    # ══════════════════════════════════════════════════

    section(f"Expense Group Trend by Year — {prop}")
    st.caption("Average cost per Full Turn by category group — high-level summary of the detailed breakdown below")

    trend_lines = p_lines[p_lines["Year"].isin(EXPENSE_YEARS)].copy()
    trend_turn_counts = (
        p_turns[p_turns["Year"].isin(EXPENSE_YEARS)]
        .groupby("Year")["Turn Key"].nunique()
        .reindex(EXPENSE_YEARS, fill_value=0)
    )

    # Map each budget category to its group
    cat_to_group = {}
    for c in CORE_LABOR:
        cat_to_group[c] = "Core Labor"
    for c in CORE_MATERIALS:
        cat_to_group[c] = "Core Materials"
    for c in OTHER_CATS:
        cat_to_group[c] = "Other"
    trend_lines["Category Group"] = trend_lines["Budget Category"].map(cat_to_group).fillna("Other")

    GROUP_ORDER = ["Core Labor", "Core Materials", "Other"]
    GROUP_COLORS = {"Core Labor": "#2563eb", "Core Materials": "#10b981", "Other": "#7c3aed"}

    grp_year_spend = (
        trend_lines.groupby(["Category Group", "Year"])["Invoice Amount"]
        .sum().reset_index(name="total_spend")
    )
    grp_year_spend["n_turns"] = grp_year_spend["Year"].map(trend_turn_counts).fillna(0)
    grp_year_spend["avg_per_turn"] = grp_year_spend.apply(
        lambda r: r["total_spend"] / r["n_turns"] if r["n_turns"] > 0 else 0, axis=1
    )

    grp_trend_pivot = grp_year_spend.pivot_table(
        index="Category Group", columns="Year",
        values="avg_per_turn", fill_value=0
    ).reindex(columns=EXPENSE_YEARS, fill_value=0).reindex(GROUP_ORDER)
    grp_trend_pivot = grp_trend_pivot.fillna(0)

    # Totals for pinned row
    grp_totals = grp_trend_pivot.sum()

    col1, col2 = st.columns([2, 3])
    with col1:
        # Add Total row
        grp_trend_pivot.loc["Total"] = grp_totals
        grp_trend_display = grp_trend_pivot.copy()
        grp_trend_display.columns = EXPENSE_YEAR_LABELS
        for col in grp_trend_display.columns:
            grp_trend_display[col] = grp_trend_display[col].apply(fmt)
        st.dataframe(
            grp_trend_display.style.set_properties(
                subset=pd.IndexSlice["Total", :],
                **{"background-color": "#f1f5f9", "font-weight": "700", "border-top": "2px solid #cbd5e1"},
            ),
            use_container_width=True,
        )

    with col2:
        fig_grp = go.Figure()
        for grp in GROUP_ORDER:
            row = grp_trend_pivot.loc[grp] if grp in grp_trend_pivot.index else pd.Series(0, index=EXPENSE_YEARS)
            fig_grp.add_trace(go.Bar(
                x=EXPENSE_YEAR_LABELS,
                y=[row.get(y, 0) for y in EXPENSE_YEARS],
                name=grp,
                marker_color=GROUP_COLORS.get(grp, "#94a3b8"),
                hovertemplate=f"{grp}<br>%{{x}}: $%{{y:,.0f}}<extra></extra>",
            ))
        fig_grp.update_layout(
            template=CHART_TEMPLATE, barmode="stack",
            xaxis=dict(title=""), yaxis=dict(title="Avg Cost per Turn ($)"),
            legend=dict(orientation="h", y=-0.15, font=dict(size=11)),
            margin=dict(t=10, b=50, l=10, r=10), height=340,
        )
        st.plotly_chart(fig_grp, use_container_width=True)

    # ══════════════════════════════════════════════════
    # DETAILED EXPENSE ANALYSIS BY BUDGET CATEGORY
    # ══════════════════════════════════════════════════
    section(f"Expense Analysis by Budget Category — {prop}")
    st.caption(f"Average cost per Full Turn by budget category ({EXPENSE_YEARS[0]} – {expense_year_label(EXPENSE_YEARS[-1])})")

    # Pre-compute: spend by budget category and year
    cat_year_spend = (
        trend_lines.groupby(["Budget Category", "Year"])["Invoice Amount"]
        .sum().reset_index(name="total_spend")
    )
    cat_year_spend["n_turns"] = cat_year_spend["Year"].map(trend_turn_counts).fillna(0)
    cat_year_spend["avg_per_turn"] = cat_year_spend.apply(
        lambda r: r["total_spend"] / r["n_turns"] if r["n_turns"] > 0 else 0, axis=1
    )

    prop_yoy_pairs = [(2023, 2024), (2024, 2025), (2025, 2026)]
    render_category_table("Core Labor (Avg per Turn)", CORE_LABOR, cat_year_spend,
                          yoy_pairs=prop_yoy_pairs)
    st.markdown("")
    render_category_table("Core Materials (Avg per Turn)", CORE_MATERIALS, cat_year_spend,
                          yoy_pairs=prop_yoy_pairs)
    st.markdown("")
    render_category_table("Other Categories (Avg per Turn)", OTHER_CATS, cat_year_spend,
                          yoy_pairs=prop_yoy_pairs)

    # ══════════════════════════════════════════════════
    # RECENT TURNS BY FLOOR PLAN (drill-down)
    # ══════════════════════════════════════════════════
    section(f"Recent Turns — {prop}")
    floor_plans = sorted(p_turns["Floor Plan"].dropna().unique())
    fp_options = ["All Floor Plans"] + floor_plans
    selected_fp = st.selectbox("Filter by Floor Plan", fp_options, key="prop_fp")

    if selected_fp == "All Floor Plans":
        fp_turns = p_turns.copy()
        fp_lines = p_lines.copy()
    else:
        fp_turns = p_turns[p_turns["Floor Plan"] == selected_fp].copy()
        fp_lines = p_lines[p_lines["Floor Plan"] == selected_fp].copy()

    fp_label = selected_fp if selected_fp != "All Floor Plans" else "All Floor Plans"

    section(f"Last 5 Completed Full Turns — {fp_label}")

    last5 = fp_turns.sort_values("completion_date", ascending=False).head(5).copy()

    if len(last5) == 0:
        st.info(f"No completed Full Turns found for {fp_label}.")
    else:
        last5["Move-Out"] = last5["Move-Out Date"].dt.strftime("%b %d, %Y").fillna("—")

        # Category Cost Comparison table — category × turn matrix
        fp_comp_keys = last5["Turn Key"].tolist()
        fp_comp_lines = fp_lines[fp_lines["Turn Key"].isin(fp_comp_keys)].copy()
        fp_comp_info = last5[["Turn Key", "Unit Label", "Move-Out Date", "Floor Plan"]].copy()
        fp_comp_info["Turn Type"] = "Full Turn"
        fp_comp_html = render_floor_plan_comparison_table(fp_comp_lines, fp_comp_info)
        st.markdown(
            f'<div class="scope-card scope-card-history">'
            f'<p class="scope-title">Category Cost Comparison — Last 5 Full Turns</p>'
            f'<p class="scope-subtitle">Compare category-level costs across the most recent {len(last5)} Full Turns'
            f'{" for " + selected_fp + " units" if selected_fp != "All Floor Plans" else ""} at {prop}</p>'
            f'{fp_comp_html}'
            f'</div>',
            unsafe_allow_html=True,
        )

        # Detail expanders — invoice-level deep dive
        for _, t in last5.iterrows():
            with st.expander(f"Detail — {t['Unit Label']} — {t['Move-Out']}"):
                items = fp_lines[fp_lines["Turn Key"] == t["Turn Key"]].sort_values("Invoice Date")
                d = items[["Vendor Name", "Budget Category", "Cost Type",
                           "Invoice Amount", "Invoice Date", "Line Item Notes"]].copy()
                d["Invoice Amount"] = d["Invoice Amount"].apply(lambda x: fmt(x, 2))
                d["Invoice Date"] = d["Invoice Date"].dt.strftime("%b %d, %Y").fillna("—")
                d["Line Item Notes"] = d["Line Item Notes"].fillna("")
                st.dataframe(d, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════
    # CLOSING NARRATIVE — PROPERTY SUMMARY
    # ══════════════════════════════════════════════════
    avg_cost = p_turns["total_cost"].mean()
    port_avg_all = ft_turns["total_cost"].mean()
    vs_port = ((avg_cost - port_avg_all) / port_avg_all * 100) if port_avg_all > 0 else 0

    # Find this property's most expensive category
    prop_cat_totals = p_lines.groupby("Budget Category")["Invoice Amount"].sum()
    top_cat = prop_cat_totals.idxmax() if len(prop_cat_totals) > 0 else "—"
    top_cat_pct = (prop_cat_totals.max() / prop_cat_totals.sum() * 100) if prop_cat_totals.sum() > 0 else 0

    # Property YoY trend
    p24 = p_turns[p_turns["Year"] == 2024]["total_cost"]
    p25 = p_turns[p_turns["Year"] == 2025]["total_cost"]
    prop_yoy = ((p25.mean() - p24.mean()) / p24.mean() * 100) if len(p24) >= 2 and len(p25) >= 2 and p24.mean() > 0 else None

    position = "above" if vs_port > 0 else "below"
    insight_text = (
        f"<strong>{prop}</strong> averages <strong>{fmt(avg_cost)}</strong>/turn — "
        f"<strong>{abs(vs_port):.1f}%</strong> {position} the portfolio average. "
        f"<strong>{top_cat}</strong> is the largest expense category at <strong>{top_cat_pct:.0f}%</strong> of spend — "
        f"{'negotiate better rates or standardize scope for this category to reduce costs.' if vs_port > 10 else 'this is in line with portfolio norms.'}"
    )
    if prop_yoy is not None:
        insight_text += (
            f" Year-over-year costs {'increased' if prop_yoy > 0 else 'decreased'} <strong>{abs(prop_yoy):.1f}%</strong> "
            f"from 2024 to 2025{'— investigate what changed in scope or vendor pricing.' if prop_yoy > 15 else '.'}"
        )
    insight(insight_text)

    footer()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 2: PORTFOLIO OVERVIEW
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif view == "2 — Portfolio Overview":
    banner("Portfolio Overview", "Cross-property cost comparison and turn volume — 2016 through 2025")

    SUMMARY_YEARS = YEARS

    # KPIs
    recent = ft_turns[ft_turns["Year"].isin(SUMMARY_YEARS)]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Full Turns (2016–25)", f"{len(recent):,}")
    c2.metric("Total Spend", fmt(recent["total_cost"].sum()))
    c3.metric("Portfolio Avg Cost", fmt(recent["total_cost"].mean()))
    c4.metric("Properties Active", f"{recent['Property ID'].nunique()}")

    section("Average Full Turn Cost by Property & Year")

    # Avg cost matrix
    avg_matrix = (
        recent.groupby(["Property Name", "Year"])["total_cost"]
        .mean().unstack(fill_value=0)
        .reindex(columns=SUMMARY_YEARS, fill_value=0)
    )

    # Count matrix
    count_matrix = (
        recent.groupby(["Property Name", "Year"]).size()
        .unstack(fill_value=0)
        .reindex(columns=SUMMARY_YEARS, fill_value=0)
    )

    # YoY change
    avg_matrix["2024 → 2025"] = avg_matrix.apply(
        lambda r: ((r[2025] - r[2024]) / r[2024] * 100) if r[2024] > 0 and r[2025] > 0 else np.nan,
        axis=1,
    )

    # Trailing 5-year avg (far-right column)
    avg_matrix["Avg (5-Yr)"] = avg_matrix[TREND_YEARS].replace(0, np.nan).mean(axis=1)

    avg_matrix = avg_matrix.loc[sorted(avg_matrix.index, key=lambda n: _PROP_RANK.get(n, 999))]

    # Portfolio total row
    portfolio_row = {}
    for y in SUMMARY_YEARS:
        yr_data = recent[recent["Year"] == y]["total_cost"]
        portfolio_row[y] = yr_data.mean() if len(yr_data) else 0
    portfolio_row["2024 → 2025"] = (
        ((portfolio_row[2025] - portfolio_row[2024]) / portfolio_row[2024] * 100)
        if portfolio_row[2024] > 0 and portfolio_row[2025] > 0 else np.nan
    )
    recent_5yr = recent[recent["Year"].isin(TREND_YEARS)]
    portfolio_row["Avg (5-Yr)"] = recent_5yr["total_cost"].mean() if len(recent_5yr) else 0
    avg_matrix.loc["PORTFOLIO AVG"] = portfolio_row

    tab_avg, tab_vol = st.tabs(["Average Cost per Turn", "Turn Volume"])

    with tab_avg:
        display = avg_matrix.copy()
        for y in SUMMARY_YEARS:
            display[y] = display[y].apply(lambda x: fmt(x) if x > 0 else "—")
        display["2024 → 2025"] = display["2024 → 2025"].apply(
            lambda x: pct(x) if pd.notna(x) else "—"
        )
        display["Avg (5-Yr)"] = display["Avg (5-Yr)"].apply(lambda x: fmt(x) if x > 0 else "—")
        # Reorder so Avg (5-Yr) is last
        col_order = [c for c in display.columns if c not in ["2024 → 2025", "Avg (5-Yr)"]] + ["2024 → 2025", "Avg (5-Yr)"]
        display = display[col_order]
        st.dataframe(
            display.style.set_properties(
                subset=["Avg (5-Yr)"],
                **{"background-color": "#eff6ff", "font-weight": "700", "color": "#1e40af"},
            ),
            use_container_width=True,
            height=560,
        )

        # Narrative
        prop_only = avg_matrix.drop("PORTFOLIO AVG")
        port_avg = avg_matrix.loc["PORTFOLIO AVG", "Avg (5-Yr)"]
        yoy_chg = avg_matrix.loc["PORTFOLIO AVG", "2024 → 2025"]

        if len(prop_only) >= 2:
            top_prop = prop_only.sort_values("Avg (5-Yr)", ascending=False).index[0]
            top_avg = prop_only.loc[top_prop, "Avg (5-Yr)"]
            low_candidates = prop_only[prop_only["Avg (5-Yr)"] > 0].sort_values("Avg (5-Yr)", ascending=True)
            low_prop = low_candidates.index[0] if len(low_candidates) else top_prop
            low_avg = prop_only.loc[low_prop, "Avg (5-Yr)"]
            spread = top_avg - low_avg
            # Count properties above/below avg
            above_avg = len(prop_only[prop_only["Avg (5-Yr)"] > port_avg])
            below_avg = len(prop_only[prop_only["Avg (5-Yr)"] <= port_avg])
            # Properties with YoY increase
            yoy_col = "2024 → 2025"
            rising = prop_only[(prop_only[yoy_col].notna()) & (prop_only[yoy_col] > 0)]
            insight(
                f"The <strong>{fmt(spread)}</strong> gap between <strong>{top_prop}</strong> "
                f"(<strong>{fmt(top_avg)}</strong>/turn) and <strong>{low_prop}</strong> "
                f"(<strong>{fmt(low_avg)}</strong>/turn) highlights inconsistent renovation standards. "
                f"<strong>{above_avg}</strong> properties run above the portfolio average of <strong>{fmt(port_avg)}</strong>. "
                f"{'<strong>' + str(len(rising)) + '</strong> properties saw cost increases from 2024 to 2025 — drill into Property Summary for those to identify drivers.' if len(rising) > 0 else 'All properties held or reduced costs from 2024 to 2025.'}"
            )
        else:
            insight(
                f"Portfolio-wide average Full Turn cost is <strong>{fmt(port_avg)}</strong>. "
                f"Year-over-year, portfolio costs moved <strong>{pct(yoy_chg)}</strong> from 2024 to 2025."
            )

    with tab_vol:
        count_matrix["Total"] = count_matrix.sum(axis=1)
        count_matrix = count_matrix.loc[sorted(count_matrix.index, key=lambda n: _PROP_RANK.get(n, 999))]
        count_matrix.loc["PORTFOLIO TOTAL"] = count_matrix.sum()
        count_matrix = count_matrix.astype(int)
        st.dataframe(count_matrix, use_container_width=True, height=560)

        # Volume insight — actionable
        vol_total = int(count_matrix.loc["PORTFOLIO TOTAL", "Total"])
        prop_only_vol = count_matrix.drop("PORTFOLIO TOTAL")
        busiest_prop = prop_only_vol["Total"].idxmax() if len(prop_only_vol) > 0 else "—"
        busiest_count = int(prop_only_vol["Total"].max()) if len(prop_only_vol) > 0 else 0
        busiest_pct = (busiest_count / vol_total * 100) if vol_total > 0 else 0
        # Recent trend
        vol_2024 = int(count_matrix.loc["PORTFOLIO TOTAL", 2024]) if 2024 in count_matrix.columns else 0
        vol_2025 = int(count_matrix.loc["PORTFOLIO TOTAL", 2025]) if 2025 in count_matrix.columns else 0
        vol_chg = vol_2025 - vol_2024
        # Lowest volume property
        lowest_prop = prop_only_vol["Total"].idxmin() if len(prop_only_vol) > 0 else "—"
        lowest_count = int(prop_only_vol["Total"].min()) if len(prop_only_vol) > 0 else 0
        insight(
            f"<strong>{busiest_prop}</strong> accounts for <strong>{busiest_pct:.0f}%</strong> of all turns "
            f"(<strong>{busiest_count}</strong> of {vol_total}) — prioritize standardizing scope and pricing there. "
            f"Turn volume went from <strong>{vol_2024}</strong> in 2024 to <strong>{vol_2025}</strong> in 2025 "
            f"(<strong>{vol_chg:+d}</strong> turns). "
            f"<strong>{lowest_prop}</strong> has the fewest turns (<strong>{lowest_count}</strong>) — "
            f"{'low volume means less data for benchmarking; combine with similar-sized properties for comparison.' if lowest_count < 10 else 'sufficient data for reliable benchmarking.'}"
        )

    # ══════════════════════════════════════════════════
    # AVERAGE COST BY PROPERTY & FLOOR PLAN (Trailing 5 Years)
    # ══════════════════════════════════════════════════
    section(f"Average Full Turn Cost by Property & Floor Plan — {TREND_YEARS[0]} through {TREND_YEARS[-1]}")
    st.caption("Average cost per Full Turn by property and floor plan (trailing 5 years only)")

    fp_recent = ft_turns[ft_turns["Year"].isin(TREND_YEARS)]
    fp_matrix = (
        fp_recent.groupby(["Property Name", "Floor Plan"])["total_cost"]
        .mean().unstack(fill_value=0)
    )
    # Sort columns by floor plan name (natural alpha)
    fp_matrix = fp_matrix.reindex(columns=sorted(fp_matrix.columns))
    # Property avg (exclude zeros)
    fp_matrix["Property Avg"] = fp_matrix[sorted(fp_matrix.columns.drop("Property Avg", errors="ignore"))].replace(0, np.nan).mean(axis=1)
    # Sort rows by PROPERTY_ORDER
    fp_matrix = fp_matrix.loc[sorted(fp_matrix.index, key=lambda n: _PROP_RANK.get(n, 999))]
    # Portfolio avg row
    fp_cols = [c for c in fp_matrix.columns if c != "Property Avg"]
    port_fp_row = {}
    for fp in fp_cols:
        fp_vals = fp_matrix[fp].replace(0, np.nan).dropna()
        port_fp_row[fp] = fp_vals.mean() if len(fp_vals) else 0
    port_fp_row["Property Avg"] = fp_recent["total_cost"].mean()
    fp_matrix.loc["PORTFOLIO AVG"] = port_fp_row

    # Format for display
    fp_display = fp_matrix.copy()
    for col in fp_display.columns:
        fp_display[col] = fp_display[col].apply(lambda x: fmt(x) if x > 0 else "—")

    st.dataframe(
        fp_display.style.set_properties(
            subset=["Property Avg"],
            **{"background-color": "#eff6ff", "font-weight": "700", "color": "#1e40af"},
        ),
        use_container_width=True,
        height=560,
    )

    # ══════════════════════════════════════════════════
    # BUDGET CATEGORY TRENDS (Avg per Turn)
    # ══════════════════════════════════════════════════
    section(f"Expense Analysis by Budget Category — {TREND_YEARS[0]} through {TREND_YEARS[-1]}")
    st.caption(f"Average cost per Full Turn by budget category ({TREND_YEARS[0]}–{TREND_YEARS[-1]})")

    # Data prep — category-level avg per turn by year
    ft_trend_po = ft_lines[ft_lines["Year"].isin(TREND_YEARS)].copy()
    trend_turn_counts_po = (
        ft_turns[ft_turns["Year"].isin(TREND_YEARS)]
        .groupby("Year")["Turn Key"].nunique()
        .reindex(TREND_YEARS, fill_value=0)
    )
    cat_year_spend_po = (
        ft_trend_po.groupby(["Budget Category", "Year"])["Invoice Amount"]
        .sum().reset_index(name="total_spend")
    )
    cat_year_spend_po["n_turns"] = cat_year_spend_po["Year"].map(trend_turn_counts_po).fillna(0)
    cat_year_spend_po["avg_per_turn"] = cat_year_spend_po.apply(
        lambda r: r["total_spend"] / r["n_turns"] if r["n_turns"] > 0 else 0, axis=1
    )
    TREND_YEAR_LABELS_PO = [str(y) for y in TREND_YEARS]

    render_category_table("Core Labor (Avg per Turn)", CORE_LABOR, cat_year_spend_po,
                          years=TREND_YEARS, year_labels=TREND_YEAR_LABELS_PO)
    st.markdown("")
    render_category_table("Core Materials (Avg per Turn)", CORE_MATERIALS, cat_year_spend_po,
                          years=TREND_YEARS, year_labels=TREND_YEAR_LABELS_PO)
    st.markdown("")
    render_category_table("Other Categories (Avg per Turn)", OTHER_CATS, cat_year_spend_po,
                          years=TREND_YEARS, year_labels=TREND_YEAR_LABELS_PO)

    footer()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 4: UNIT SEARCH
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif view == "4 — Unit Search":
    banner("Unit Search", "Unit-level turn history, pre-scoping checklist, and projected turn cost")

    # Use all-types turn summary for this view
    all_turns = build_turn_summary(_df_all)

    col1, col2 = st.columns(2)
    with col1:
        _all_unit_props = list(_df_all["Property Name"].unique())
        _all_unit_props_ordered = [p for p in PROPERTIES if p in _all_unit_props] + sorted(set(_all_unit_props) - set(PROPERTIES))
        prop_choice = st.selectbox("Property", _all_unit_props_ordered)
    with col2:
        prop_units = sorted(_df_all[_df_all["Property Name"] == prop_choice]["Unit Label"].unique())
        unit_choice = st.selectbox("Unit", prop_units)

    unit_ts = all_turns[
        (all_turns["Property Name"] == prop_choice) & (all_turns["Unit Label"] == unit_choice)
    ].sort_values("Move-Out Date", ascending=False)

    unit_df = _df_all[
        (_df_all["Property Name"] == prop_choice) & (_df_all["Unit Label"] == unit_choice)
    ]

    if len(unit_ts) == 0:
        st.info("No turn history found for this unit.")
    else:
        sample = unit_df.iloc[0]
        bed = f"{sample['Bedrooms']:.0f}" if pd.notna(sample['Bedrooms']) else "—"
        bath = f"{sample['Bathrooms']:.0f}" if pd.notna(sample['Bathrooms']) else "—"
        st.markdown(
            f"**{prop_choice} — Unit {unit_choice}** &nbsp;|&nbsp; "
            f"Floor Plan: `{sample['Floor Plan']}` &nbsp;|&nbsp; "
            f"Bed/Bath: `{bed} / {bath}` &nbsp;|&nbsp; "
            f"ID: `{sample['UID']}`"
        )

        most_recent = unit_ts.iloc[0]
        most_recent_date = most_recent["Move-Out Date"]
        most_recent_type = most_recent["Turn Type"]

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Turns", len(unit_ts))
        c2.metric("Lifetime Spend", fmt(unit_ts["total_cost"].sum()))
        c3.metric("Avg Cost / Turn", fmt(unit_ts["total_cost"].mean()))
        c4.metric("Most Recent", f"{most_recent_type} ({most_recent_date.strftime('%b %Y') if pd.notna(most_recent_date) else '—'})")

        section("Turn History")
        for _, turn in unit_ts.iterrows():
            with st.expander(
                f"**{turn['Turn Type']}** — {turn['Move-Out Date'].strftime('%b %d, %Y') if pd.notna(turn['Move-Out Date']) else '—'} — "
                f"{fmt(turn['total_cost'])} ({turn['line_items']} invoices)"
            ):
                items = unit_df[unit_df["Turn Key"] == turn["Turn Key"]].sort_values("Invoice Date")
                disp = items[["Invoice Date", "Vendor Name", "Budget Category",
                              "Cost Type", "Invoice Amount", "Line Item Notes"]].copy()
                disp["Invoice Date"] = disp["Invoice Date"].dt.strftime("%b %d, %Y").fillna("—")
                disp["Invoice Amount"] = disp["Invoice Amount"].apply(lambda x: fmt(x, 2))
                disp["Line Item Notes"] = disp["Line Item Notes"].fillna("")
                st.dataframe(disp, use_container_width=True, hide_index=True)

        # ══════════════════════════════════════════════════
        # PRE-SCOPING — UNIT WORK HISTORY
        # ══════════════════════════════════════════════════
        unit_ft = unit_ts[unit_ts["Turn Type"] == "Full Turn"]

        # Compute last Full Turn cost (needed for projected scope comparison)
        total_last_ft = 0
        if len(unit_ft) > 0:
            last_ft = unit_ft.sort_values("Move-Out Date", ascending=False).iloc[0]
            last_ft_items = unit_df[unit_df["Turn Key"] == last_ft["Turn Key"]]
            total_last_ft = last_ft_items["Invoice Amount"].sum()

        section("Pre-Scoping — Unit Work History")

        # Multi-year category table — all turn types
        history_html, hist_years, hist_totals = render_scope_history_table(unit_df, unit_ts)
        ALL_CATS = CORE_LABOR + CORE_MATERIALS + OTHER_CATS
        total_cats = len(ALL_CATS)
        # Count categories ever touched
        work_pivot = unit_df.copy()
        work_pivot["Year"] = work_pivot["Move-Out Date"].dt.year
        cat_ever = work_pivot.groupby("Budget Category")["Invoice Amount"].sum()
        cats_touched = sum(1 for c in ALL_CATS if cat_ever.get(c, 0) > 0)

        st.markdown(
            f'<div class="scope-card scope-card-history">'
            f'<p class="scope-title">Unit Work History — {cats_touched} of {total_cats} Categories Touched</p>'
            f'<p class="scope-subtitle">All turns on this unit (Full Turn, Make Ready, Partial Turn) — most recent first</p>'
            f'{history_html}'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Export Work History (Excel / PDF) ──
        if hist_years:
            # Build export DataFrame: category × year pivot
            work_exp = unit_df.copy()
            work_exp["Year"] = work_exp["Move-Out Date"].dt.year
            work_pivot_exp = work_exp.pivot_table(
                index="Budget Category", columns="Year",
                values="Invoice Amount", aggfunc="sum", fill_value=0,
            ).reindex(columns=hist_years, fill_value=0)
            ALL_CATS_EXP = CORE_LABOR + CORE_MATERIALS + OTHER_CATS
            group_labels = {}
            for c in CORE_LABOR:
                group_labels[c] = "Core Labor"
            for c in CORE_MATERIALS:
                group_labels[c] = "Core Materials"
            for c in OTHER_CATS:
                group_labels[c] = "Other"
            rows = []
            for c in ALL_CATS_EXP:
                row = {"Category": c, "Group": group_labels.get(c, "")}
                for y in hist_years:
                    row[str(int(y))] = work_pivot_exp.loc[c, y] if c in work_pivot_exp.index else 0
                rows.append(row)
            hist_export_df = pd.DataFrame(rows)
            # Add total row
            total_row_exp = {"Category": "Total", "Group": ""}
            for y in hist_years:
                total_row_exp[str(int(y))] = hist_totals.get(y, 0)
            hist_export_df = pd.concat([hist_export_df, pd.DataFrame([total_row_exp])], ignore_index=True)

            # Excel — single sheet with header context
            hist_xl_buf = io.BytesIO()
            with pd.ExcelWriter(hist_xl_buf, engine="openpyxl") as writer:
                hist_export_df.to_excel(writer, sheet_name="Work History", index=False, startrow=2)
                ws = writer.sheets["Work History"]
                fp_xl = unit_df.iloc[0]["Floor Plan"] if len(unit_df) > 0 else ""
                ws.cell(row=1, column=1, value=f"{prop_choice}  |  Unit {unit_choice}  |  {fp_xl}  |  {cats_touched} of {total_cats} categories  |  {pd.Timestamp.now().strftime('%B %d, %Y')}")
                ws.cell(row=1, column=1).font = Font(bold=True, size=11)
            hist_xl_buf.seek(0)
            hist_xl_name = f"{prop_choice}_{unit_choice}_work_history_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"

            # PDF
            hist_pdf = FPDF(orientation="L", unit="mm", format="A4")
            hist_pdf.set_auto_page_break(auto=True, margin=15)
            hist_pdf.add_page()
            hist_pdf.set_font("Helvetica", "B", 14)
            hist_pdf.cell(0, 10, _pdf(f"Unit Work History - {prop_choice} - Unit {unit_choice}"), ln=True)
            hist_pdf.set_font("Helvetica", "", 9)
            fp_label_pdf = unit_df.iloc[0]["Floor Plan"] if len(unit_df) > 0 else ""
            hist_pdf.cell(0, 6, _pdf(f"Floor Plan: {fp_label_pdf}  |  {cats_touched} of {total_cats} categories touched  |  Generated {pd.Timestamp.now().strftime('%B %d, %Y')}"), ln=True)
            hist_pdf.ln(4)

            # Table header
            year_strs = [str(int(y)) for y in hist_years]
            cat_col_w = 60
            grp_col_w = 35
            yr_col_w = min(28, (297 - 20 - cat_col_w - grp_col_w) / max(len(hist_years), 1))
            hist_pdf.set_font("Helvetica", "B", 8)
            hist_pdf.cell(cat_col_w, 7, "Category", border=1)
            hist_pdf.cell(grp_col_w, 7, "Group", border=1)
            for ys in year_strs:
                hist_pdf.cell(yr_col_w, 7, ys, border=1, align="R")
            hist_pdf.ln()

            # Table rows — skip categories with all-zero spend
            hist_pdf.set_font("Helvetica", "", 8)
            for _, r in hist_export_df.iterrows():
                is_total = r["Category"] == "Total"
                row_sum = sum(r.get(ys, 0) for ys in year_strs)
                if not is_total and row_sum == 0:
                    continue
                if is_total:
                    hist_pdf.set_font("Helvetica", "B", 8)
                hist_pdf.cell(cat_col_w, 6, _pdf(r["Category"]), border=1)
                hist_pdf.cell(grp_col_w, 6, _pdf(r["Group"]), border=1)
                for ys in year_strs:
                    val = r.get(ys, 0)
                    hist_pdf.cell(yr_col_w, 6, _pdf(fmt(val)) if val != 0 else "-", border=1, align="R")
                hist_pdf.ln()
                if is_total:
                    hist_pdf.set_font("Helvetica", "", 8)

            hist_pdf_buf = io.BytesIO(hist_pdf.output())
            hist_pdf_name = f"{prop_choice}_{unit_choice}_work_history_{pd.Timestamp.now().strftime('%Y%m%d')}.pdf"

            dl1, dl2, _ = st.columns([1, 1, 4])
            with dl1:
                st.download_button("📥 Excel", hist_xl_buf, file_name=hist_xl_name,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="hist_xl")
            with dl2:
                st.download_button("📥 PDF", hist_pdf_buf, file_name=hist_pdf_name,
                                   mime="application/pdf", key="hist_pdf")

        # Vendor summary — all turns
        vendor_brkdn = (
            unit_df.groupby("Vendor Name")["Invoice Amount"]
            .sum().reset_index(name="Amount")
            .sort_values("Amount", ascending=False)
        )
        vendor_brkdn["Amount"] = vendor_brkdn["Amount"].apply(lambda x: fmt(x, 2))
        vendor_brkdn.columns = ["Vendor", "Total"]
        with st.expander("Vendor Breakdown — All Turns"):
            st.dataframe(vendor_brkdn, use_container_width=True, hide_index=True)

        # ══════════════════════════════════════════════════
        # PROJECTED TURN COST
        # ══════════════════════════════════════════════════
        section("Projected Turn Cost — Recommended Scope & Budget")

        unit_floor_plan = unit_df.iloc[0]["Floor Plan"] if len(unit_df) > 0 else None
        has_prior_ft = len(unit_ft) > 0

        # Determine projected turn type
        if has_prior_ft:
            proj_type = "Make Ready"
        else:
            proj_type = "Full Turn"

        # Build comps based on turn type:
        #   Make Ready → last 5 at the property (any floor plan, most recent pricing)
        #   Full Turn  → same floor plan + last 2 years (floor plan matters for scope)
        if proj_type == "Make Ready":
            # Last 5 Make Readys at this property (any floor plan)
            mr_turn_totals = (
                _df_all[
                    (_df_all["Property Name"] == prop_choice)
                    & (_df_all["Turn Type"] == "Make Ready")
                ].groupby("Turn Key").agg(
                    move_out=("Move-Out Date", "first"),
                ).reset_index()
                .sort_values("move_out", ascending=False)
                .head(5)
            )
            comp_keys = mr_turn_totals["Turn Key"].tolist()
            comp_lines = _df_all[_df_all["Turn Key"].isin(comp_keys)].copy()
            comp_turn_keys = len(comp_keys)
            comp_desc = f"last {comp_turn_keys} Make Readys at {prop_choice}"
            st.caption(
                f"This unit has a prior Full Turn — projecting a **Make Ready** scope. "
                f"Based on the {comp_desc}."
            )
        else:
            cutoff_2yr = pd.Timestamp.now() - pd.DateOffset(years=2)
            comp_lines = _df_all[
                (_df_all["Property Name"] == prop_choice)
                & (_df_all["Floor Plan"] == unit_floor_plan)
                & (_df_all["Turn Type"] == "Full Turn")
                & (_df_all["Move-Out Date"] >= cutoff_2yr)
            ].copy()
            comp_turn_keys = comp_lines["Turn Key"].nunique()
            comp_desc = f"{comp_turn_keys} comparable Full Turns for {unit_floor_plan} units at {prop_choice} (last 2 years)"
            st.caption(
                f"No prior Full Turn on record — projecting a **Full Turn** scope. "
                f"Based on {comp_desc}."
            )

        if comp_turn_keys >= 1:
            # Category-level avg per turn
            comp_cat = (
                comp_lines.groupby("Budget Category")["Invoice Amount"]
                .sum().reset_index(name="total_spend")
            )
            comp_cat["Avg per Turn"] = comp_cat["total_spend"] / comp_turn_keys

            # Category frequency filter — for Make Readies, only include categories
            # that appeared in ≥50% of comp turns. This removes one-off items like
            # Windows and Scrape Ceiling that are typically Full-Turn-only scope.
            if proj_type == "Make Ready" and comp_turn_keys >= 3:
                cat_freq = (
                    comp_lines.groupby("Budget Category")["Turn Key"]
                    .nunique().reset_index(name="turn_count")
                )
                cat_freq["frequency"] = cat_freq["turn_count"] / comp_turn_keys
                common_cats = cat_freq[cat_freq["frequency"] >= 0.50]["Budget Category"].tolist()
                excluded = comp_cat[~comp_cat["Budget Category"].isin(common_cats)]
                comp_cat = comp_cat[comp_cat["Budget Category"].isin(common_cats)].copy()
                if len(excluded) > 0:
                    excluded_names = excluded["Budget Category"].tolist()
                    excluded_total = excluded["Avg per Turn"].sum()
                    st.caption(
                        f"ℹ️ Excluded from projection: **{', '.join(excluded_names)}** "
                        f"(appeared in <50% of comps, typically one-time Full Turn items — "
                        f"${excluded_total:,.0f}/turn removed)."
                    )

            projected_total = comp_cat["Avg per Turn"].sum()
            proj_amounts = comp_cat.set_index("Budget Category")["Avg per Turn"]

            # ── Recency adjustment — exclude categories done on THIS unit within last 2 years ──
            RECENCY_CUTOFF_YEARS = 2
            recency_cutoff = pd.Timestamp.now() - pd.DateOffset(years=RECENCY_CUTOFF_YEARS)
            recent_unit_work = unit_df[unit_df["Move-Out Date"] >= recency_cutoff].copy()
            recently_done_agg = (
                recent_unit_work.groupby("Budget Category")["Invoice Amount"]
                .sum().reset_index(name="spend")
            )
            recently_done_agg = recently_done_agg[recently_done_agg["spend"] > 0]
            recently_done_cats = set(recently_done_agg["Budget Category"])
            last_done_year = (
                recent_unit_work[recent_unit_work["Invoice Amount"] > 0]
                .groupby("Budget Category")["Move-Out Date"].max().dt.year.to_dict()
            )

            # Build adjusted amounts — zero out categories recently done on this unit
            excluded_recent = {}
            for cat in recently_done_cats:
                if cat in proj_amounts.index and proj_amounts[cat] > 0:
                    excluded_recent[cat] = proj_amounts[cat]

            adjusted_amounts = proj_amounts.copy()
            for cat in excluded_recent:
                adjusted_amounts[cat] = 0
            adjusted_total = adjusted_amounts[adjusted_amounts > 0].sum()
            savings_from_recent = sum(excluded_recent.values())

            # KPIs
            c1, c2, c3 = st.columns(3)
            c1.metric(
                f"Projected {proj_type} Cost", fmt(adjusted_total),
                f"-{fmt(savings_from_recent)} recent work" if savings_from_recent > 0 else None,
            )
            c2.metric("Based on Comps", f"{comp_turn_keys} {proj_type}s")
            c3.metric("Recently Done", f"{len(excluded_recent)} categories" if excluded_recent else "None")

            if excluded_recent:
                excluded_names = [f"{c} ({last_done_year.get(c, '?')})" for c in excluded_recent]
                st.caption(
                    f"**Adjusted for recent work on this unit:** {', '.join(excluded_names)} "
                    f"completed within the last {RECENCY_CUTOFF_YEARS} years — "
                    f"comp average of ${savings_from_recent:,.0f} removed. "
                    f"Unadjusted baseline: {fmt(projected_total)}."
                )

            # Visual checklist — projected scope with comp columns
            expected_count = sum(1 for c in CORE_LABOR + CORE_MATERIALS + OTHER_CATS
                                 if proj_amounts.get(c, 0) > 0 and c not in excluded_recent)
            total_cats = len(CORE_LABOR) + len(CORE_MATERIALS) + len(OTHER_CATS)

            # Build comp info for table columns
            comp_info_df = comp_lines.groupby("Turn Key").agg(
                Unit_Label=("Unit Label", "first"),
                Move_Out_Date=("Move-Out Date", "first"),
                Turn_Type=("Turn Type", "first"),
            ).reset_index()
            comp_info_df.columns = ["Turn Key", "Unit Label", "Move-Out Date", "Turn Type"]

            proj_checklist_html = render_projected_scope_table(
                proj_amounts, excluded_recent, last_done_year, amount_label="Projected Cost",
                comp_lines=comp_lines, comp_info=comp_info_df,
            )
            st.markdown(
                f'<div class="scope-card scope-card-projected">'
                f'<p class="scope-title">Recommended Scope — {proj_type} &nbsp;|&nbsp; {expected_count} of {total_cats} Categories</p>'
                f'<p class="scope-subtitle">Based on {comp_desc} — adjusted for recent work on this unit</p>'
                f'{proj_checklist_html}'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── Export buttons — download scope as Excel or PDF ──
            scope_rows = []
            for group_label, group_cats in [("Core Labor", CORE_LABOR), ("Core Materials", CORE_MATERIALS), ("Other", OTHER_CATS)]:
                for c in group_cats:
                    amt = adjusted_amounts.get(c, 0) if c in adjusted_amounts.index else 0
                    if c in excluded_recent:
                        scope_rows.append({"Category": c, "Group": group_label, "Projected Cost": excluded_recent[c], "Status": f"Recently Done ({last_done_year.get(c, '')})"})
                    elif amt > 0:
                        scope_rows.append({"Category": c, "Group": group_label, "Projected Cost": amt, "Status": "Included"})
                    else:
                        scope_rows.append({"Category": c, "Group": group_label, "Projected Cost": 0, "Status": "Not Expected"})
            scope_export_df = pd.DataFrame(scope_rows)
            today_str = datetime.now().strftime("%Y-%m-%d")

            exp_c1, exp_c2, _ = st.columns([1, 1, 4])
            # Excel download
            excel_buf = io.BytesIO()
            with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                # Summary sheet
                summary_data = pd.DataFrame([{
                    "Property": prop_choice,
                    "Unit": unit_choice,
                    "Floor Plan": unit_floor_plan,
                    "Projection Type": proj_type,
                    "Projected Total": adjusted_total,
                    "Based On": comp_desc,
                    "Date Generated": today_str,
                }])
                summary_data.to_excel(writer, sheet_name="Summary", index=False)
                # Scope detail sheet
                scope_export_df.to_excel(writer, sheet_name="Scope Detail", index=False)
            excel_buf.seek(0)
            exp_c1.download_button(
                "📥 Excel",
                data=excel_buf.getvalue(),
                file_name=f"{prop_choice}_{unit_choice}_scope_{today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # PDF download
            pdf = FPDF(orientation="L", unit="mm", format="A4")
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, _pdf(f"Recommended Scope - {proj_type}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, _pdf(f"{prop_choice}  |  Unit {unit_choice}  |  {unit_floor_plan}  |  Generated {today_str}"), new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 6, _pdf(f"Based on {comp_desc}"), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(4)
            # Table header
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(241, 245, 249)
            pdf.cell(90, 7, "Category", border=1, fill=True)
            pdf.cell(45, 7, "Group", border=1, fill=True)
            pdf.cell(40, 7, "Projected Cost", border=1, fill=True, align="R")
            pdf.cell(50, 7, "Status", border=1, fill=True)
            pdf.ln()
            # Table rows
            pdf.set_font("Helvetica", "", 9)
            for _, r in scope_export_df.iterrows():
                if r["Projected Cost"] > 0 or r["Status"] != "Not Expected":
                    pdf.cell(90, 6, _pdf(r["Category"]), border=1)
                    pdf.cell(45, 6, _pdf(r["Group"]), border=1)
                    pdf.cell(40, 6, _pdf(fmt(r["Projected Cost"])), border=1, align="R")
                    pdf.cell(50, 6, _pdf(r["Status"]), border=1)
                    pdf.ln()
            # Total row
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(135, 7, "PROJECTED TOTAL", border=1, fill=True)
            pdf.cell(40, 7, _pdf(fmt(adjusted_total)), border=1, fill=True, align="R")
            pdf.cell(50, 7, "", border=1, fill=True)
            pdf.ln()

            pdf_buf = io.BytesIO()
            pdf.output(pdf_buf)
            pdf_buf.seek(0)
            exp_c2.download_button(
                "📥 PDF",
                data=pdf_buf.getvalue(),
                file_name=f"{prop_choice}_{unit_choice}_scope_{today_str}.pdf",
                mime="application/pdf",
            )

            # Budget guidance insight
            top_adj = adjusted_amounts[adjusted_amounts > 0].sort_values(ascending=False)
            top_proj_name = top_adj.index[0] if len(top_adj) > 0 else "N/A"
            top_proj_val = top_adj.iloc[0] if len(top_adj) > 0 else 0
            top_proj_pct = top_proj_val / adjusted_total * 100 if adjusted_total > 0 else 0

            if adjusted_total == 0 and excluded_recent:
                # All projected categories were recently done on this unit
                insight(
                    f"<strong>Budget Guidance:</strong> All comparable {proj_type} categories have been "
                    f"completed on this unit within the last {RECENCY_CUTOFF_YEARS} years — "
                    f"<strong>{len(excluded_recent)}</strong> categories totaling "
                    f"<strong>{fmt(savings_from_recent)}</strong> (comp average) excluded. "
                    f"Unadjusted baseline was <strong>{fmt(projected_total)}</strong>. "
                    f"<strong>Action:</strong> Minimal scope expected — confirm unit condition with "
                    f"a physical walkthrough before authorizing any work orders."
                )
            elif has_prior_ft:
                mr_ratio = adjusted_total / total_last_ft * 100 if total_last_ft > 0 else 0
                recent_note = (
                    f" <strong>{len(excluded_recent)}</strong> categories were recently completed on this unit "
                    f"and excluded from the projection (saving <strong>{fmt(savings_from_recent)}</strong>)."
                    if excluded_recent else ""
                )
                insight(
                    f"<strong>Budget Guidance:</strong> Projected <strong>{proj_type}</strong> cost is "
                    f"<strong>{fmt(adjusted_total)}</strong> — <strong>{mr_ratio:.0f}%</strong> of the "
                    f"last Full Turn (<strong>{fmt(total_last_ft)}</strong>).{recent_note} "
                    f"The largest remaining line item is <strong>{top_proj_name}</strong> at "
                    f"<strong>{fmt(top_proj_val)}</strong> ({top_proj_pct:.0f}% of total). "
                    f"<strong>Action:</strong> Get competitive bids on {top_proj_name} to keep this "
                    f"{proj_type} under <strong>{fmt(adjusted_total * 0.9)}</strong> (10% savings target)."
                )
            else:
                port_avg_type = ft_turns["total_cost"].mean() if proj_type == "Full Turn" else 0
                vs_port = ""
                if port_avg_type > 0:
                    diff_pct = (adjusted_total - port_avg_type) / port_avg_type * 100
                    position = "above" if diff_pct > 0 else "below"
                    vs_port = (
                        f" This is <strong>{abs(diff_pct):.0f}%</strong> {position} the portfolio "
                        f"Full Turn average of <strong>{fmt(port_avg_type)}</strong>."
                    )
                insight(
                    f"<strong>Budget Guidance:</strong> Projected <strong>{proj_type}</strong> cost is "
                    f"<strong>{fmt(adjusted_total)}</strong> based on <strong>{comp_desc}</strong>.{vs_port} "
                    f"The largest projected line item is <strong>{top_proj_name}</strong> at "
                    f"<strong>{fmt(top_proj_val)}</strong> ({top_proj_pct:.0f}% of total). "
                    f"<strong>Action:</strong> Prioritize scope validation on the top 3 categories — "
                    f"confirm necessity before approving work orders to avoid scope creep."
                )

        else:
            # Fallback: no comps found at all
            if proj_type == "Full Turn":
                # Try property-wide Full Turns (any floor plan, last 2 years)
                cutoff_2yr = pd.Timestamp.now() - pd.DateOffset(years=2)
                fallback_lines = _df_all[
                    (_df_all["Property Name"] == prop_choice)
                    & (_df_all["Turn Type"] == "Full Turn")
                    & (_df_all["Move-Out Date"] >= cutoff_2yr)
                ].copy()
                fallback_turns = fallback_lines["Turn Key"].nunique()
                if fallback_turns > 0:
                    fallback_total = fallback_lines["Invoice Amount"].sum() / fallback_turns
                    st.info(
                        f"No Full Turn comps for {unit_floor_plan} at {prop_choice} in the last 2 years. "
                        f"Using property-wide Full Turn average: **{fmt(fallback_total)}** "
                        f"(based on {fallback_turns} turns)."
                    )
                else:
                    st.info(f"No recent Full Turn data available at {prop_choice} for projection.")
            else:
                st.info(f"No Make Ready data available at {prop_choice} for projection.")

    footer()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 1: EXECUTIVE SUMMARY
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif view == "1 — Executive Summary":
    banner("Executive Summary", "Portfolio-wide performance, benchmarking, and risk flags — 2016 through 2025")

    # ── Compute core metrics ──
    curr_year = ft_turns[ft_turns["Year"] == 2025]
    prev_year = ft_turns[ft_turns["Year"] == 2024]

    portfolio_avg = ft_turns["total_cost"].mean()
    curr_avg = curr_year["total_cost"].mean() if len(curr_year) else 0
    prev_avg = prev_year["total_cost"].mean() if len(prev_year) else 0
    yoy_delta = ((curr_avg - prev_avg) / prev_avg * 100) if prev_avg > 0 else np.nan

    curr_vol = len(curr_year)
    prev_vol = len(prev_year)
    vol_delta = ((curr_vol - prev_vol) / prev_vol * 100) if prev_vol > 0 else np.nan

    dur_curr = curr_year["Duration"].dropna()
    dur_prev = prev_year["Duration"].dropna()
    dur_now = dur_curr.median() if len(dur_curr) else np.nan
    dur_then = dur_prev.median() if len(dur_prev) else np.nan

    # ── Top-Line KPIs ──
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("2025 Full Turns", f"{curr_vol:,}", f"{vol_delta:+.0f}% vs '24" if pd.notna(vol_delta) else "—")
    c2.metric("2025 Avg Cost", fmt(curr_avg), (pct(yoy_delta) + " vs '24") if pd.notna(yoy_delta) else "—")
    c3.metric("Portfolio Avg (All)", fmt(portfolio_avg))
    c4.metric("2025 Total Spend", fmt(curr_year["total_cost"].sum()))
    c5.metric("Median Duration", f"{dur_now:.0f} days" if pd.notna(dur_now) else "—",
              f"{dur_now - dur_then:+.0f} days vs '24" if (pd.notna(dur_now) and pd.notna(dur_then)) else "")
    c6.metric("Active Properties", f"{curr_year['Property ID'].nunique()}" if len(curr_year) else "0")

    # ━━ Section 1: Cost Trajectory ━━
    section("Cost Trajectory — Are We Getting More Efficient?")

    yearly_stats = ft_turns[ft_turns["Year"].isin(YEARS)].groupby("Year").agg(
        avg_cost=("total_cost", "mean"),
        median_cost=("total_cost", "median"),
        total_spend=("total_cost", "sum"),
        turn_count=("Turn Key", "count"),
    ).reset_index()

    col1, col2 = st.columns([3, 2])
    with col1:
        fig_traj = go.Figure()
        fig_traj.add_trace(go.Scatter(
            x=yearly_stats["Year"], y=yearly_stats["avg_cost"],
            name="Avg Cost", mode="lines+markers",
            line=dict(color="#2563eb", width=3), marker=dict(size=9),
            hovertemplate="Avg: $%{y:,.0f}<extra></extra>",
        ))
        fig_traj.add_trace(go.Scatter(
            x=yearly_stats["Year"], y=yearly_stats["median_cost"],
            name="Median Cost", mode="lines+markers",
            line=dict(color="#10b981", width=2, dash="dash"), marker=dict(size=7),
            hovertemplate="Median: $%{y:,.0f}<extra></extra>",
        ))
        fig_traj.update_layout(
            template=CHART_TEMPLATE,
            xaxis=dict(dtick=1, title=""), yaxis=dict(title="Cost per Full Turn ($)"),
            legend=dict(orientation="h", y=-0.12), margin=dict(t=10, b=50), height=380,
            hovermode="x unified",
        )
        st.plotly_chart(fig_traj, use_container_width=True)

    with col2:
        # Year-over-year change table
        yearly_stats["YoY Change"] = yearly_stats["avg_cost"].pct_change() * 100
        ys_display = yearly_stats[["Year", "turn_count", "avg_cost", "median_cost", "total_spend", "YoY Change"]].copy()
        ys_display["avg_cost"] = ys_display["avg_cost"].apply(fmt)
        ys_display["median_cost"] = ys_display["median_cost"].apply(fmt)
        ys_display["total_spend"] = ys_display["total_spend"].apply(fmt)
        ys_display["YoY Change"] = ys_display["YoY Change"].apply(lambda x: pct(x) if pd.notna(x) else "—")
        ys_display["Year"] = ys_display["Year"].astype(int).astype(str)
        ys_display.columns = ["Year", "Turns", "Avg Cost", "Median", "Total Spend", "YoY Δ"]
        st.dataframe(ys_display, use_container_width=True, hide_index=True, height=380)

    # Cost efficiency narrative
    if len(yearly_stats) >= 2:
        last_3 = yearly_stats[yearly_stats["Year"].isin([2023, 2024, 2025])]
        if len(last_3) >= 2:
            recent_trend = ((last_3.iloc[-1]["avg_cost"] - last_3.iloc[0]["avg_cost"]) / last_3.iloc[0]["avg_cost"]) * 100 if last_3.iloc[0]["avg_cost"] > 0 else 0
            trend_word = "rising" if recent_trend > 5 else ("declining" if recent_trend < -5 else "stable")
        else:
            recent_trend = 0
            trend_word = "stable"
        avg_vs_med = curr_avg - (curr_year["total_cost"].median() if len(curr_year) else 0)
        skew_note = (
            f" The gap between average and median in 2025 is <strong>{fmt(abs(avg_vs_med))}</strong>, "
            f"indicating {'a few high-cost turns are pulling the average up — review those outliers for scope creep' if avg_vs_med > 1000 else 'consistent turn costs across the portfolio'}."
        ) if len(curr_year) >= 3 else ""
        insight(
            f"Costs are <strong>{trend_word}</strong> over the last 3 years (<strong>{pct(recent_trend)}</strong> from 2023 to 2025). "
            f"{'Focus on controlling scope at properties with rising costs — see benchmarking below.' if recent_trend > 5 else 'Cost discipline is holding — maintain current vendor agreements and scope standards.' if recent_trend <= 5 and recent_trend >= -5 else 'Cost reductions are working — document what changed and apply across the portfolio.'}"
            f"{skew_note}"
        )

    # ━━ Section 2: Property Benchmarking ━━
    section("Property Benchmarking — Cost Efficiency Ranking")

    prop_bench = ft_turns[ft_turns["Year"].isin(YEARS)].groupby("Property Name").agg(
        turns=("Turn Key", "count"),
        avg_cost=("total_cost", "mean"),
        median_cost=("total_cost", "median"),
        total_spend=("total_cost", "sum"),
        med_duration=("Duration", "median"),
    ).reset_index().sort_values("avg_cost", ascending=False)

    # Add rank (1 = highest avg cost)
    prop_bench["Rank"] = range(1, len(prop_bench) + 1)
    prop_bench["vs Portfolio"] = ((prop_bench["avg_cost"] - portfolio_avg) / portfolio_avg * 100) if portfolio_avg > 0 else 0

    col1, col2 = st.columns([3, 2])
    with col1:
        # Chart: lowest cost at top, highest at bottom (ascending for horizontal bar)
        chart_data = prop_bench.sort_values("avg_cost", ascending=True)
        fig_bench = px.bar(
            chart_data,
            y="Property Name", x="avg_cost", orientation="h",
            text=chart_data["avg_cost"].apply(fmt),
            template=CHART_TEMPLATE,
            color="avg_cost",
            color_continuous_scale=["#10b981", "#f59e0b", "#dc2626"],
        )
        fig_bench.update_traces(textposition="outside")
        fig_bench.add_vline(x=portfolio_avg, line_dash="dash", line_color="#2563eb",
                            annotation_text=f"Portfolio Avg: {fmt(portfolio_avg)}",
                            annotation_position="top right", annotation_font_size=11)
        fig_bench.update_layout(
            margin=dict(t=20, b=10, l=10, r=100), height=480,
            xaxis_title="Avg Full Turn Cost ($)", yaxis_title="",
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig_bench, use_container_width=True)

    with col2:
        pb_display = prop_bench[["Rank", "Property Name", "turns", "avg_cost", "vs Portfolio", "med_duration"]].copy()
        pb_display["avg_cost"] = pb_display["avg_cost"].apply(fmt)
        pb_display["vs Portfolio"] = pb_display["vs Portfolio"].apply(lambda x: pct(x))
        pb_display["med_duration"] = pb_display["med_duration"].apply(lambda x: f"{x:.0f} days" if pd.notna(x) else "—")
        pb_display.columns = ["#", "Property", "Turns", "Avg Cost", "vs Portfolio", "Med Duration"]
        st.dataframe(pb_display, use_container_width=True, hide_index=True, height=480)

    # Identify best and worst performers with actionable gap analysis
    if len(prop_bench) >= 2:
        best = prop_bench.sort_values("avg_cost").iloc[0]
        worst = prop_bench.sort_values("avg_cost").iloc[-1]
        cost_gap = worst["avg_cost"] - best["avg_cost"]
        # Calculate potential savings if worst performed at portfolio avg
        worst_excess_per_turn = worst["avg_cost"] - portfolio_avg
        worst_potential_savings = worst_excess_per_turn * worst["turns"]
        insight(
            f"<strong>{worst['Property Name']}</strong> averages <strong>{fmt(cost_gap)}</strong> more per turn "
            f"than <strong>{best['Property Name']}</strong>. "
            f"If {worst['Property Name']} operated at the portfolio average, it would have saved approximately "
            f"<strong>{fmt(worst_potential_savings)}</strong> over <strong>{int(worst['turns'])}</strong> turns. "
            f"<strong>Action:</strong> Compare vendor rates, scope of work, and material choices between these "
            f"two properties to identify where the cost gap originates."
        )

    # ━━ Section 3: Vendor Concentration ━━
    section("Vendor Concentration — Risk & Spend Distribution")

    vendor_data = ft_lines[ft_lines["Year"].isin(YEARS)].groupby("Vendor Name").agg(
        total_spend=("Invoice Amount", "sum"),
        invoices=("Invoice Amount", "count"),
        properties=("Property Name", "nunique"),
        avg_invoice=("Invoice Amount", "mean"),
    ).reset_index().sort_values("total_spend", ascending=False)

    total_vendor_spend = vendor_data["total_spend"].sum()
    vendor_data["Share"] = (vendor_data["total_spend"] / total_vendor_spend * 100) if total_vendor_spend > 0 else 0
    vendor_data["Cumulative"] = vendor_data["Share"].cumsum()

    top_n = 10
    top_vendors = vendor_data.head(top_n)
    top_share = top_vendors["Share"].sum()

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Vendors", f"{len(vendor_data):,}")
    c2.metric(f"Top {top_n} Share", f"{top_share:.1f}%")
    c3.metric("Single-Property Vendors", f"{len(vendor_data[vendor_data['properties'] == 1]):,}")

    col1, col2 = st.columns([3, 2])
    with col1:
        fig_vendor = px.bar(
            top_vendors, y="Vendor Name", x="total_spend", orientation="h",
            text=top_vendors["total_spend"].apply(fmt),
            template=CHART_TEMPLATE, color_discrete_sequence=["#6366f1"],
        )
        fig_vendor.update_traces(textposition="outside")
        fig_vendor.update_layout(
            margin=dict(t=10, b=10, l=10, r=100), height=400,
            xaxis_title="Total Spend ($)", yaxis_title="",
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(fig_vendor, use_container_width=True)

    with col2:
        v_display = top_vendors[["Vendor Name", "total_spend", "Share", "invoices", "properties"]].copy()
        v_display["total_spend"] = v_display["total_spend"].apply(fmt)
        v_display["Share"] = v_display["Share"].apply(lambda x: f"{x:.1f}%")
        v_display.columns = ["Vendor", "Total Spend", "Portfolio %", "Invoices", "Properties"]
        st.dataframe(v_display, use_container_width=True, hide_index=True, height=400)

    # Vendor narrative — actionable
    top_vendor = top_vendors.iloc[0]["Vendor Name"] if len(top_vendors) > 0 else "—"
    top_vendor_share = top_vendors.iloc[0]["Share"] if len(top_vendors) > 0 else 0
    top_vendor_spend = top_vendors.iloc[0]["total_spend"] if len(top_vendors) > 0 else 0
    single_prop_count = len(vendor_data[vendor_data["properties"] == 1])
    single_prop_spend = vendor_data[vendor_data["properties"] == 1]["total_spend"].sum()
    single_prop_pct = (single_prop_spend / total_vendor_spend * 100) if total_vendor_spend > 0 else 0
    insight(
        f"<strong>{top_vendor}</strong> handles <strong>{top_vendor_share:.1f}%</strong> of all spend "
        f"(<strong>{fmt(top_vendor_spend)}</strong>). "
        f"<strong>{single_prop_count}</strong> vendors work at only one property, accounting for "
        f"<strong>{single_prop_pct:.0f}%</strong> of spend — these are candidates for competitive bidding "
        f"or consolidation with multi-property vendors who may offer volume discounts. "
        f"<strong>Action:</strong> {'The top 10 vendors control ' + f'{top_share:.0f}% of spend — negotiate volume pricing or add backup vendors to reduce dependency.' if top_share > 60 else 'Vendor spread is healthy — continue monitoring for pricing consistency across properties.'}"
    )

    # ━━ Section 4: Capital Forecast ━━
    section("Capital Forecast — Projected Annual Full Turn Spend")

    # Use 3-year trailing average for projection
    hist_years = [2023, 2024, 2025]
    hist_data = ft_turns[ft_turns["Year"].isin(hist_years)].groupby("Year").agg(
        turns=("Turn Key", "count"),
        total=("total_cost", "sum"),
        avg_cost=("total_cost", "mean"),
    ).reset_index()

    if len(hist_data) >= 2:
        avg_turns = hist_data["turns"].mean()
        avg_cost_trend = hist_data["avg_cost"].mean()

        # Cost trend (linear)
        cost_slope = np.polyfit(hist_data["Year"], hist_data["avg_cost"], 1)
        projected_cost_2026 = max(np.polyval(cost_slope, 2026), 0)

        vol_slope = np.polyfit(hist_data["Year"], hist_data["turns"], 1)
        projected_vol_2026 = max(np.polyval(vol_slope, 2026), 0)

        forecast_spend = projected_vol_2026 * projected_cost_2026

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("3-Year Avg Turns/Year", f"{avg_turns:.0f}")
        c2.metric("3-Year Avg Cost", fmt(avg_cost_trend))
        c3.metric("2026 Proj. Avg Cost", fmt(projected_cost_2026))
        c4.metric("2026 Proj. Spend", fmt(forecast_spend))

        # Forecast chart
        all_years_data = ft_turns[ft_turns["Year"].isin(YEARS)].groupby("Year").agg(
            total=("total_cost", "sum"),
        ).reset_index()

        fig_fc = go.Figure()
        fig_fc.add_trace(go.Bar(
            x=all_years_data["Year"], y=all_years_data["total"],
            name="Actual", marker_color="#2563eb",
            hovertemplate="Actual: $%{y:,.0f}<extra></extra>",
        ))
        fig_fc.add_trace(go.Bar(
            x=[2026], y=[forecast_spend],
            name="Projected", marker_color="#94a3b8", opacity=0.6,
            hovertemplate="Projected: $%{y:,.0f}<extra></extra>",
        ))
        fig_fc.update_layout(
            template=CHART_TEMPLATE, barmode="group",
            xaxis=dict(dtick=1, title=""), yaxis=dict(title="Annual Full Turn Spend ($)"),
            legend=dict(orientation="h", y=-0.12), margin=dict(t=10, b=50), height=380,
        )
        st.plotly_chart(fig_fc, use_container_width=True)

        # Compare projected vs actual recent spend
        actual_2025 = hist_data[hist_data["Year"] == 2025]["total"].iloc[0] if 2025 in hist_data["Year"].values else 0
        forecast_vs_2025 = ((forecast_spend - actual_2025) / actual_2025 * 100) if actual_2025 > 0 else 0
        insight(
            f"2026 projection: <strong>{projected_vol_2026:.0f}</strong> Full Turns at "
            f"<strong>{fmt(projected_cost_2026)}</strong>/turn = <strong>{fmt(forecast_spend)}</strong> total capital. "
            f"That is <strong>{pct(forecast_vs_2025)}</strong> vs 2025 actual spend of <strong>{fmt(actual_2025)}</strong>. "
            f"<strong>Action:</strong> {'Budget accordingly for increased capital needs — consider pre-negotiating vendor rates before volume increases.' if forecast_vs_2025 > 10 else 'Spend is projected to hold steady — lock in current vendor pricing to protect margins.' if forecast_vs_2025 >= -10 else 'Declining volume creates an opportunity to be more selective on scope and vendor quality.'}"
        )

    # ━━ Section 5: Key Risk Flags ━━
    section("Risk Flags — Items Requiring Executive Attention")

    risk_items = []

    # 1. Properties with rising costs (2024 → 2025)
    for prop in PROPERTIES:
        p24 = ft_turns[(ft_turns["Property Name"] == prop) & (ft_turns["Year"] == 2024)]["total_cost"]
        p25 = ft_turns[(ft_turns["Property Name"] == prop) & (ft_turns["Year"] == 2025)]["total_cost"]
        if len(p24) >= 2 and len(p25) >= 2 and p24.mean() > 0:
            chg = ((p25.mean() - p24.mean()) / p24.mean()) * 100
            if chg > 20:
                risk_items.append({
                    "Risk": "Cost Escalation",
                    "Detail": f"{prop}: avg cost up {chg:.0f}% YoY ({fmt(p24.mean())} → {fmt(p25.mean())})",
                    "Severity": "High" if chg > 40 else "Medium",
                })

    # 2. Properties with slow duration
    for prop in PROPERTIES:
        p_dur = ft_turns[(ft_turns["Property Name"] == prop) & (ft_turns["Year"] == 2025)]["Duration"].dropna()
        if len(p_dur) >= 3 and p_dur.median() > 60:
            risk_items.append({
                "Risk": "Slow Velocity",
                "Detail": f"{prop}: median turn duration {p_dur.median():.0f} days (>60d threshold)",
                "Severity": "Medium",
            })

    # 3. Category inflation (per-turn spend, all 17 raw categories)
    for cat_name in MATERIALS_CATS + LABOR_CATS:
        cat24_turns = ft_lines[(ft_lines["Year"] == 2024) & (ft_lines["Budget Category"] == cat_name)].groupby("Turn Key")["Invoice Amount"].sum()
        cat25_turns = ft_lines[(ft_lines["Year"] == 2025) & (ft_lines["Budget Category"] == cat_name)].groupby("Turn Key")["Invoice Amount"].sum()
        if len(cat24_turns) >= 5 and len(cat25_turns) >= 5:
            avg24, avg25 = cat24_turns.mean(), cat25_turns.mean()
            chg = ((avg25 - avg24) / avg24) * 100 if avg24 > 0 else 0
            if chg > 25 and (avg25 - avg24) > 100:
                risk_items.append({
                    "Risk": "Category Inflation",
                    "Detail": f"{cat_name}: avg per-turn spend up {chg:.0f}% YoY ({fmt(avg24)} → {fmt(avg25)})",
                    "Severity": "High" if chg > 50 else "Medium",
                })

    # 4. High-frequency units (>4 Full Turns all-time)
    freq_units = ft_turns.groupby(["Property Name", "Unit Label"])["Turn Key"].nunique().reset_index(name="turns")
    chronic = freq_units[freq_units["turns"] >= 5]
    if len(chronic) > 0:
        risk_items.append({
            "Risk": "Chronic Vacancy",
            "Detail": f"{len(chronic)} units with 5+ Full Turns — potential workmanship or tenant screening issues",
            "Severity": "Medium",
        })

    if risk_items:
        risk_df = pd.DataFrame(risk_items)
        # Sort: High first
        severity_order = {"High": 0, "Medium": 1, "Low": 2}
        risk_df["_sort"] = risk_df["Severity"].map(severity_order)
        risk_df = risk_df.sort_values("_sort").drop("_sort", axis=1)

        for _, r in risk_df.iterrows():
            sev_class = "risk-flag-high" if r["Severity"] == "High" else "risk-flag-medium"
            icon = "🔴" if r["Severity"] == "High" else "🟡"
            st.markdown(
                f'<div class="risk-flag {sev_class}">'
                f'<strong>{icon} {r["Risk"]}</strong> — {r["Detail"]}</div>',
                unsafe_allow_html=True,
            )

        st.markdown("")  # spacer
        high_count = len([r for r in risk_items if r["Severity"] == "High"])
        med_count = len([r for r in risk_items if r["Severity"] == "Medium"])
        escalation_count = len([r for r in risk_items if r["Risk"] == "Cost Escalation"])
        cat_inflation_count = len([r for r in risk_items if r["Risk"] == "Category Inflation"])
        priority_text = []
        if escalation_count > 0:
            priority_text.append(f"Investigate the <strong>{escalation_count}</strong> cost escalation {'flags' if escalation_count > 1 else 'flag'} first — pull invoices at those properties and compare scope vs prior year")
        if cat_inflation_count > 0:
            priority_text.append(f"Review the <strong>{cat_inflation_count}</strong> category inflation {'items' if cat_inflation_count > 1 else 'item'} — request updated vendor bids or negotiate fixed pricing")
        action_text = ". ".join(priority_text) + "." if priority_text else "Monitor current metrics through the next quarter."
        insight(
            f"<strong>{high_count}</strong> high-severity and <strong>{med_count}</strong> medium-severity flags. "
            f"<strong>Priority actions:</strong> {action_text}"
        )
    else:
        st.success("No risk flags identified — portfolio metrics are within normal ranges.")

    footer()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 5: RENT ROLL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif view == "5 — Rent Roll":
    banner("Rent Roll", "Current occupancy with historical turn activity")

    # ── Rent roll file registry ──
    _RR_DIR = Path(__file__).parent / "Rent Rolls"
    _RR_FILES = {
        "Monterey Park": "Rent Roll - MontereyPark.xlsx",
        "Woodman": "Rent Roll - Woodman.xlsx",
        "Collins": "Rent Roll - Collins.xlsx",
        "51 at the Village": "Rent Roll - 51Village.xlsx",
        "Lindley": "Rent Roll - Lindley.xlsx",
        "El Rancho": "Rent Roll - ElRancho.xlsx",
        "Alta Vista": "Rent Roll - Alta Vista.xlsx",
        "Roscoe": "Rent Roll - Roscoe.xlsx",
        "Woodbridge": "Rent Roll - Woodbridge.xlsx",
        "Darby": "Rent Roll - Darby.xlsx",
        "Dickens": "Rent Roll - Dickens.xlsx",
        "Fruitland": "Rent Roll - Fruitland.xlsx",
        "Garfield": "Rent Roll - Garfield.xlsx",
        "12756 Moorpark": "Rent Roll - 12756Moorpark.xlsx",
        "12800 Moorpark": "Rent Roll - 12800Moorpark.xlsx",
    }

    # ── Load rent roll ──
    @st.cache_data
    def load_rent_roll(path):
        """Load and clean a rent roll Excel file."""
        rr = pd.read_excel(path, header=8)
        rr = rr.dropna(how="all")
        rr = rr[rr["Unit"].notna() & ~rr["Unit"].astype(str).str.contains("Units|Total", na=False)]
        rr = rr[~rr["Unit"].astype(str).str.contains("LLC|Properties|, LP", na=False)]
        rr["Market Rent"] = pd.to_numeric(rr["Market Rent"], errors="coerce")
        rr["Rent"] = pd.to_numeric(rr["Rent"], errors="coerce")
        rr["Move-in"] = pd.to_datetime(rr["Move-in"], errors="coerce")
        return rr

    # ── Unit mapping: rent roll → turn data ──
    def _norm_unit(u):
        """Normalize a unit identifier: strip known suffixes (HUD/MGR/BC),
        remove formatting hyphens ('12-A' → '12A'), normalize leading zeros ('01' → '1')."""
        s = str(u).strip()
        # Strip known suffixes (space- or hyphen-delimited)
        for suffix in (" HUD", " MGR", " BC", "-MGR", " ASST. MGR"):
            if s.upper().endswith(suffix.upper()):
                s = s[:len(s) - len(suffix)].strip()
                break
        # Remove formatting hyphens in unit IDs ("12-A" → "12A")
        if "-" in s:
            s = s.replace("-", "")
        # Normalize leading zeros for purely numeric units
        if s.isdigit() and len(s) > 1:
            s = str(int(s))
        return s

    def _normalize_mp_bldg(addr):
        """Normalize a Monterey Park building address to match turn data Building Code format.
        E.g. '511 W Pomona Blvd' → '511 Pomona', '2425 Hendricks Ave' → '2425 Hendericks'."""
        parts = addr.strip().split()
        number = parts[0]
        rest = parts[1:]
        # Skip directional prefixes (W, S, N, E)
        if rest and rest[0].upper() in ("W", "S", "N", "E"):
            rest = rest[1:]
        # Drop address suffixes
        suffixes = {"Ave", "Blvd", "Dr", "St", "Ln", "Way", "Ct",
                    "AVE", "BLVD", "DR", "ST", "LN", "WAY", "CT"}
        street_parts = [p for p in rest if p not in suffixes]
        street = " ".join(street_parts)
        # Known spelling correction (rent roll vs turn data)
        if street == "Hendricks":
            street = "Hendericks"
        return f"{number} {street}"

    def rr_to_turn_key(prop_name, rr_unit_str):
        """Map a rent roll unit ID to the compound key used in turn data.
        Returns 'BuildingCode|UnitNumber' for properties with building codes,
        or just 'UnitNumber' for direct-match properties."""
        u = str(rr_unit_str).strip()

        if prop_name == "Collins":
            # "39-01" → "18339 Collins|1", "39-21-MGR" → "18339 Collins|21"
            parts = u.split("-")
            if len(parts) >= 2:
                prefix = parts[0]
                unit_num = _norm_unit(parts[1])
                bldg = {"39": "18339 Collins", "47": "18347 Collins"}.get(prefix)
                if bldg:
                    return f"{bldg}|{unit_num}"
            return _norm_unit(u)

        elif prop_name == "Monterey Park":
            # "505 Pomona, Unit A" → "505 Pomona|A"
            # "511 W Pomona Blvd, Unit C" → "511 Pomona|C"
            # "617 Pomona, Unit B - ASST. MGR" → "617 Pomona|B"
            # "515 W Pomona Blvd" (no unit) → "515 Pomona"
            if ", Unit " in u:
                bldg_raw, unit_raw = u.rsplit(", Unit ", 1)
                unit_letter = unit_raw.split()[0].rstrip(" -")
            elif " Unit " in u:
                idx = u.index(" Unit ")
                bldg_raw = u[:idx]
                unit_raw = u[idx + 6:]
                unit_letter = unit_raw.split()[0].rstrip(" -")
            else:
                bldg_raw = u
                unit_letter = None
            norm_bldg = _normalize_mp_bldg(bldg_raw)
            return f"{norm_bldg}|{unit_letter}" if unit_letter else norm_bldg

        else:
            # All other properties: strip suffixes, normalize leading zeros
            return _norm_unit(u)

    # ── Build turn history columns for a property ──
    TURN_ABBR = {"Full Turn": "FT", "Make Ready": "MR", "Partial Turn": "PT"}

    @st.cache_data
    def build_turn_history(prop_name, _df_all):
        """Return a dict: compound_key → dict of year → 'FT - $26,000' string.
        If multiple turns in the same year, entries are joined with ' / '."""
        prop_lines = _df_all[_df_all["Property Name"] == prop_name].copy()
        prop_lines["Move-Out Date"] = pd.to_datetime(prop_lines["Move-Out Date"], errors="coerce")
        # Build compound key from turn data (normalize unit numbers for consistent matching)
        prop_lines["_bldg"] = prop_lines["Building Code"].fillna("").astype(str).str.strip()
        prop_lines["_unit"] = prop_lines["Unit Number"].astype(str).str.strip().apply(_norm_unit)
        prop_lines["_key"] = prop_lines.apply(
            lambda r: f"{r['_bldg']}|{r['_unit']}" if r["_bldg"] else r["_unit"], axis=1
        )
        turns = prop_lines.groupby(["_key", "Move-Out Date", "Turn Type"]).agg(
            Total_Cost=("Invoice Amount", "sum")
        ).reset_index()
        turns = turns.sort_values("Move-Out Date", ascending=False)

        history = {}
        for key, grp in turns.groupby("_key"):
            year_data = {}
            for _, r in grp.iterrows():
                abbr = TURN_ABBR.get(r["Turn Type"], "??")
                yr = r["Move-Out Date"].year
                entry = f"{abbr} - {fmt(r['Total_Cost'])}"
                if yr in year_data:
                    year_data[yr] += f" / {entry}"
                else:
                    year_data[yr] = entry
            history[key] = year_data
        return history

    @st.cache_data
    def get_ft_units(prop_name, _df_all):
        """Return set of compound keys for units that have had at least one Full Turn."""
        ft = _df_all[(_df_all["Property Name"] == prop_name) & (_df_all["Turn Type"] == "Full Turn")].copy()
        ft["_bldg"] = ft["Building Code"].fillna("").astype(str).str.strip()
        ft["_unit"] = ft["Unit Number"].astype(str).str.strip().apply(_norm_unit)
        ft["_key"] = ft.apply(
            lambda r: f"{r['_bldg']}|{r['_unit']}" if r["_bldg"] else r["_unit"], axis=1
        )
        return set(ft["_key"])

    # ── Portfolio Summary — Classic vs Full Turn ──
    section("Portfolio Summary — Classic vs. Full Turn")
    st.caption("Classic = unit with no Full Turn on record  •  Properties without rent rolls show turn data only")

    summary_rows = []
    for prop in PROPERTY_ORDER:
        rr_file = _RR_FILES.get(prop)
        rr_path = _RR_DIR / rr_file if rr_file else None
        has_rr = rr_path is not None and rr_path.exists()

        if has_rr:
            rr_data = load_rent_roll(rr_path)
            rr_units = set(rr_data["Unit"].astype(str).str.strip())
            ft_keys = get_ft_units(prop, _df_all)
            # Map rent roll unit IDs to turn data compound keys for matching
            rr_mapped = {rr_to_turn_key(prop, u) for u in rr_units}
            n_units = len(rr_units)
            n_ft = len(rr_mapped & ft_keys)
            n_classic = n_units - n_ft
            ren_pct = f"{n_ft / n_units * 100:.1f}%" if n_units else "-"
            classic_pct = f"{n_classic / n_units * 100:.1f}%" if n_units else "-"
            summary_rows.append({
                "Property": prop,
                "Units": n_units,
                "Renovated": n_ft,
                "Classic": n_classic,
                "% Renovated": ren_pct,
                "% Classic": classic_pct,
            })
        else:
            # Placeholder — no rent roll yet
            summary_rows.append({
                "Property": prop,
                "Units": "-",
                "Renovated": "-",
                "Classic": "-",
                "% Renovated": "-",
                "% Classic": "-",
            })

    summary_df = pd.DataFrame(summary_rows)

    def _highlight_pct_cols(col):
        if col.name in ("% Renovated", "% Classic"):
            return ["background-color: #f0f4f8"] * len(col)
        return [""] * len(col)

    st.dataframe(
        summary_df.style.apply(_highlight_pct_cols),
        use_container_width=True, hide_index=True,
    )

    st.markdown("---")

    # ── Property selector for unit-level detail ──
    rr_props = [p for p in PROPERTY_ORDER if p in _RR_FILES and (_RR_DIR / _RR_FILES[p]).exists()]
    if not rr_props:
        st.info("No rent roll files found.")
    else:
        prop_choice = st.selectbox("Select Property", rr_props)
        rr_path = _RR_DIR / _RR_FILES[prop_choice]
        rr = load_rent_roll(rr_path)
        turn_hist = build_turn_history(prop_choice, _df_all)
        ft_keys = get_ft_units(prop_choice, _df_all)

        # ── KPIs ──
        total_units = len(rr)
        total_market = rr["Market Rent"].sum()
        total_rent = rr["Rent"].sum()
        ltl = total_market - total_rent
        ltl_pct = ltl / total_market if total_market else 0
        rr_unit_set = set(rr["Unit"].astype(str).str.strip())
        rr_mapped = {rr_to_turn_key(prop_choice, u) for u in rr_unit_set}
        n_ft = len(rr_mapped & ft_keys)
        n_classic = total_units - n_ft

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Total Units", f"{total_units}")
        k2.metric("Monthly Rent", fmt(total_rent))
        k3.metric("Loss to Lease", fmt(ltl), f"{ltl_pct * 100:.1f}%")
        k4.metric("Full Turns", f"{n_ft}")
        k5.metric("Classic", f"{n_classic}", f"{n_classic / total_units * 100:.1f}%" if total_units else "-")

        # ── Build combined table with year columns ──
        _YEAR_COLS = [2026, 2025, 2024, 2023, 2022, 2021]

        display_rows = []
        for _, row in rr.iterrows():
            unit_key = str(row["Unit"]).strip()
            r = {
                "Unit": unit_key,
                "BD/BA": row["BD/BA"],
                "Market Rent": fmt(row["Market Rent"]) if pd.notna(row["Market Rent"]) else "",
                "Rent": fmt(row["Rent"]) if pd.notna(row["Rent"]) else "",
                "Move-in": row["Move-in"].strftime("%b %Y") if pd.notna(row["Move-in"]) else "",
            }
            mapped_key = rr_to_turn_key(prop_choice, unit_key)
            year_data = turn_hist.get(mapped_key, {})
            for yr in _YEAR_COLS:
                r[str(yr)] = year_data.get(yr, "")
            display_rows.append(r)

        display_df = pd.DataFrame(display_rows)

        section(f"Rent Roll — {prop_choice}")
        st.caption(f"{total_units} units  •  {n_ft} renovated  •  {n_classic} classic")

        def _highlight_year_cols(col):
            if col.name in [str(y) for y in _YEAR_COLS]:
                return ["background-color: #f0f4f8"] * len(col)
            return [""] * len(col)

        st.dataframe(
            display_df.style.apply(_highlight_year_cols),
            use_container_width=True, hide_index=True, height=700,
        )

        # ── Totals ──
        tc1, tc2 = st.columns(2)
        tc1.markdown(f"**Total Market Rent:** {fmt(total_market)}")
        tc2.markdown(f"**Total Rent:** {fmt(total_rent)}")

    st.markdown("---")

    # ── High-Frequency Turn Outliers ──
    section("High-Frequency Turn Outliers")
    st.caption("Units with 5+ distinct turns since 2019 — flags frequent vacancy, tenant issues, or excessive rework")

    _out_df = _df_all[pd.to_datetime(_df_all["Move-Out Date"], errors="coerce").dt.year >= 2019].copy()
    _out_df["Move-Out Date"] = pd.to_datetime(_out_df["Move-Out Date"], errors="coerce")
    _out_df["_bldg"] = _out_df["Building Code"].fillna("").astype(str).str.strip()

    # Each distinct turn = unique (Property, BuildingCode, UnitNumber, MoveOutDate, TurnType)
    _out_turns = _out_df.groupby(
        ["Property Name", "_bldg", "Unit Number", "Move-Out Date", "Turn Type"]
    ).agg(Cost=("Invoice Amount", "sum")).reset_index()

    _out_unit = _out_turns.groupby(["Property Name", "_bldg", "Unit Number"]).agg(
        Turns=("Turn Type", "count"),
        Total_Spent=("Cost", "sum"),
    ).reset_index()

    _outliers = _out_unit[_out_unit["Turns"] >= 5].sort_values(
        ["Turns", "Total_Spent"], ascending=[False, False]
    )

    if _outliers.empty:
        st.info("No units found with 5+ turns since 2019.")
    else:
        outlier_rows = []
        for _, orow in _outliers.iterrows():
            mask = (
                (_out_turns["Property Name"] == orow["Property Name"])
                & (_out_turns["_bldg"] == orow["_bldg"])
                & (_out_turns["Unit Number"] == orow["Unit Number"])
            )
            detail = _out_turns[mask].sort_values("Move-Out Date")
            summaries = []
            for _, t in detail.iterrows():
                abbr = TURN_ABBR.get(t["Turn Type"], "??")
                yr = t["Move-Out Date"].year
                summaries.append(f"{abbr} {yr} – {fmt(t['Cost'])}")

            outlier_rows.append({
                "Property": orow["Property Name"],
                "Unit": orow["Unit Number"],
                "Turns": int(orow["Turns"]),
                "Total Spent": fmt(orow["Total_Spent"]),
                "Turn History (since 2019)": "  →  ".join(summaries),
            })

        insight(
            f"<strong>{len(outlier_rows)} units</strong> across the portfolio have had "
            f"<strong>5 or more turns since 2019</strong>. "
            f"Frequent turnover drives up costs and signals potential tenant retention "
            f"or unit condition issues that warrant investigation."
        )
        st.dataframe(pd.DataFrame(outlier_rows), use_container_width=True, hide_index=True)

    footer()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 6: DATA HEALTH
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif view == "6 — Data Health":
    banner("Data Health", "Data source freshness and update compliance")

    _APP_DIR = Path(__file__).parent
    _STALE_DAYS = 90  # 3 months

    def file_health(path, label):
        """Return dict with file label, last modified, age in days, and health status."""
        if not path.exists():
            return {"Source": label, "File": path.name, "Last Updated": "Not found", "Age": "-", "Status": "\u274c Missing"}
        mtime = datetime.fromtimestamp(os.path.getmtime(path))
        age_days = (datetime.now() - mtime).days
        status = "\u2705 Current" if age_days <= _STALE_DAYS else "\ud83d\udd34 Stale"
        return {
            "Source": label,
            "File": path.name,
            "Last Updated": mtime.strftime("%b %d, %Y"),
            "Age": f"{age_days} days",
            "Status": status,
        }

    health_rows = []

    # Turn data
    turn_path = _APP_DIR / "Unit Turns - AI Clean - 2.26.2026.xlsx"
    health_rows.append(file_health(turn_path, "Unit Turn Data"))

    # Rent rolls
    _RR_DIR_H = _APP_DIR / "Rent Rolls"
    rr_files = {
        "Rent Roll — Monterey Park": "Rent Roll - MontereyPark.xlsx",
        "Rent Roll — Woodman": "Rent Roll - Woodman.xlsx",
        "Rent Roll — Collins": "Rent Roll - Collins.xlsx",
        "Rent Roll — 51 at the Village": "Rent Roll - 51Village.xlsx",
        "Rent Roll — Lindley": "Rent Roll - Lindley.xlsx",
        "Rent Roll — El Rancho": "Rent Roll - ElRancho.xlsx",
        "Rent Roll — Alta Vista": "Rent Roll - Alta Vista.xlsx",
        "Rent Roll — Roscoe": "Rent Roll - Roscoe.xlsx",
        "Rent Roll — Woodbridge": "Rent Roll - Woodbridge.xlsx",
        "Rent Roll — Darby": "Rent Roll - Darby.xlsx",
        "Rent Roll — Dickens": "Rent Roll - Dickens.xlsx",
        "Rent Roll — Fruitland": "Rent Roll - Fruitland.xlsx",
        "Rent Roll — Garfield": "Rent Roll - Garfield.xlsx",
        "Rent Roll — 12756 Moorpark": "Rent Roll - 12756Moorpark.xlsx",
        "Rent Roll — 12800 Moorpark": "Rent Roll - 12800Moorpark.xlsx",
    }
    for label, fname in rr_files.items():
        health_rows.append(file_health(_RR_DIR_H / fname, label))

    # Vacancy
    vac_path = _APP_DIR / "Vacancy" / "Unit Vacancy.xlsx"
    health_rows.append(file_health(vac_path, "Unit Vacancy"))

    health_df = pd.DataFrame(health_rows)

    section("Data Source Status")
    st.caption(f"Sources older than {_STALE_DAYS} days are flagged as stale  •  Target: update every 3 months")

    st.dataframe(health_df, use_container_width=True, hide_index=True)

    # Summary metrics
    total_sources = len(health_rows)
    current = sum(1 for r in health_rows if "\u2705" in r["Status"])
    stale = sum(1 for r in health_rows if "\ud83d\udd34" in r["Status"])
    missing = sum(1 for r in health_rows if "\u274c" in r["Status"])

    k1, k2, k3 = st.columns(3)
    k1.metric("Total Sources", f"{total_sources}")
    k2.metric("Current", f"{current}")
    k3.metric("Stale / Missing", f"{stale + missing}", delta=f"-{stale + missing}" if (stale + missing) > 0 else None)

    footer()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VIEW 7: AI DATA REVIEW
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif view == "7 — AI Data Review":
    banner("AI Data Review", "Ask questions about your portfolio data — powered by AI")

    # ── Build data context for the LLM ──
    @st.cache_data
    def build_data_context(_ft_turns, _ft_lines, _df_all):
        """Generate a compact summary of the portfolio data for the LLM system prompt."""
        lines = []
        lines.append("=== PORTFOLIO OVERVIEW ===")
        lines.append(f"Total Full Turns: {len(_ft_turns):,}")
        lines.append(f"Properties: {_ft_turns['Property Name'].nunique()}")
        lines.append(f"Unique Units: {_ft_turns['UID'].nunique()}")
        min_date = _ft_turns['Move-Out Date'].min()
        max_date = _ft_turns['Move-Out Date'].max()
        if pd.notna(min_date) and pd.notna(max_date):
            lines.append(f"Date Range: {min_date.strftime('%b %Y')} to {max_date.strftime('%b %Y')}")
        lines.append(f"Total Spend: ${_ft_turns['total_cost'].sum():,.0f}")
        lines.append(f"Avg Cost per Turn: ${_ft_turns['total_cost'].mean():,.0f}")
        lines.append(f"Median Cost per Turn: ${_ft_turns['total_cost'].median():,.0f}")
        dur = _ft_turns["Duration"].dropna()
        if len(dur) > 0:
            lines.append(f"Avg Duration: {dur.mean():.0f} days | Median: {dur.median():.0f} days")

        lines.append("\n=== PROPERTY BREAKDOWN ===")
        prop_stats = _ft_turns.groupby("Property Name").agg(
            turns=("Turn Key", "count"),
            avg_cost=("total_cost", "mean"),
            total_spend=("total_cost", "sum"),
        )
        prop_stats = prop_stats.loc[sorted(prop_stats.index, key=lambda n: _PROP_RANK.get(n, 999))]
        for prop, row in prop_stats.iterrows():
            lines.append(f"  {prop}: {row['turns']} turns, avg ${row['avg_cost']:,.0f}, total ${row['total_spend']:,.0f}")

        lines.append("\n=== YEARLY TRENDS ===")
        yearly = _ft_turns.groupby("Year").agg(
            turns=("Turn Key", "count"),
            avg_cost=("total_cost", "mean"),
            total_spend=("total_cost", "sum"),
        )
        for yr, row in yearly.iterrows():
            lines.append(f"  {int(yr)}: {row['turns']} turns, avg ${row['avg_cost']:,.0f}, total ${row['total_spend']:,.0f}")

        lines.append("\n=== TOP BUDGET CATEGORIES (by total spend) ===")
        cat_spend = _ft_lines.groupby("Budget Category")["Invoice Amount"].agg(["sum", "mean", "count"]).sort_values("sum", ascending=False)
        for cat, row in cat_spend.head(17).iterrows():
            lines.append(f"  {cat}: ${row['sum']:,.0f} total, ${row['mean']:,.0f} avg invoice, {int(row['count'])} invoices")

        lines.append("\n=== FLOOR PLAN MIX ===")
        fp_stats = _ft_turns.groupby("Floor Plan").agg(
            turns=("Turn Key", "count"),
            avg_cost=("total_cost", "mean"),
        ).sort_values("turns", ascending=False)
        for fp, row in fp_stats.head(10).iterrows():
            lines.append(f"  {fp}: {row['turns']} turns, avg ${row['avg_cost']:,.0f}")

        lines.append("\n=== VENDOR SUMMARY (Top 15) ===")
        vendor_stats = _ft_lines.groupby("Vendor Name")["Invoice Amount"].agg(["sum", "count"]).sort_values("sum", ascending=False)
        for v, row in vendor_stats.head(15).iterrows():
            lines.append(f"  {v}: ${row['sum']:,.0f} total, {int(row['count'])} invoices")

        # ── A. All Turn Types Summary ──
        lines.append("\n=== ALL TURN TYPES (not just Full Turns) ===")
        all_ts = _df_all.copy()
        all_ts["Turn Key"] = all_ts["UID"] + "|" + all_ts["Move-Out Date"].astype(str)
        turn_type_stats = all_ts.groupby("Turn Type").agg(
            total_spend=("Invoice Amount", "sum"),
            turns=("Turn Key", "nunique"),
        )
        for tt, row in turn_type_stats.iterrows():
            avg = row["total_spend"] / row["turns"] if row["turns"] > 0 else 0
            lines.append(f"  {tt}: {row['turns']} turns, avg ${avg:,.0f}, total ${row['total_spend']:,.0f}")

        # ── B. Cost Type Breakdown ──
        lines.append("\n=== COST TYPE BREAKDOWN (Full Turns) ===")
        if "Cost Type" in _ft_lines.columns:
            ft_total_spend = _ft_lines["Invoice Amount"].sum()
            ct_stats = _ft_lines.groupby("Cost Type")["Invoice Amount"].agg(["sum", "count"]).sort_values("sum", ascending=False)
            num_ft = len(_ft_turns)
            for ct, row in ct_stats.iterrows():
                ct_pct = row["sum"] / ft_total_spend * 100 if ft_total_spend > 0 else 0
                avg_per_turn = row["sum"] / num_ft if num_ft > 0 else 0
                lines.append(f"  {ct}: ${row['sum']:,.0f} ({ct_pct:.1f}% of spend), ${avg_per_turn:,.0f} avg/turn, {int(row['count'])} invoices")

        # ── C. Property × Year Matrix ──
        lines.append("\n=== PROPERTY × YEAR DETAIL (Full Turns) ===")
        py_stats = _ft_turns.groupby(["Property Name", "Year"]).agg(
            turns=("Turn Key", "count"),
            avg_cost=("total_cost", "mean"),
            total_spend=("total_cost", "sum"),
        )
        for (prop, yr), row in py_stats.iterrows():
            lines.append(f"  {prop} | {int(yr)}: {row['turns']} turns, avg ${row['avg_cost']:,.0f}, total ${row['total_spend']:,.0f}")

        # ── D. Duration by Property ──
        lines.append("\n=== DURATION BY PROPERTY (Full Turns, days) ===")
        dur_by_prop = _ft_turns.dropna(subset=["Duration"]).groupby("Property Name")["Duration"].agg(["mean", "median", "min", "max"])
        dur_by_prop = dur_by_prop.loc[sorted(dur_by_prop.index, key=lambda n: _PROP_RANK.get(n, 999))]
        for prop, row in dur_by_prop.iterrows():
            lines.append(f"  {prop}: avg {row['mean']:.0f}d, median {row['median']:.0f}d, range {row['min']:.0f}–{row['max']:.0f}d")

        # ── E. Category × Year Trends ──
        lines.append("\n=== BUDGET CATEGORY × YEAR TRENDS (Full Turns) ===")
        _ft_lines_yr = _ft_lines.copy()
        _ft_lines_yr["Year"] = _ft_lines_yr["Move-Out Date"].dt.year
        cat_yr = _ft_lines_yr.groupby(["Budget Category", "Year"])["Invoice Amount"].sum().reset_index(name="total")
        turns_per_yr = _ft_turns.groupby("Year")["Turn Key"].count()
        for (cat, yr), row_val in cat_yr.set_index(["Budget Category", "Year"]).iterrows():
            n = turns_per_yr.get(yr, 1)
            avg = row_val["total"] / n if n > 0 else 0
            lines.append(f"  {cat} | {int(yr)}: ${row_val['total']:,.0f} total, ${avg:,.0f} avg/turn")

        # ── F. Property × Category Top Spend ──
        lines.append("\n=== TOP 5 CATEGORIES BY PROPERTY (Full Turns) ===")
        prop_cat = _ft_lines.groupby(["Property Name", "Budget Category"])["Invoice Amount"].sum().reset_index(name="total")
        for prop in sorted(_ft_turns["Property Name"].unique(), key=lambda n: _PROP_RANK.get(n, 999)):
            p_cats = prop_cat[prop_cat["Property Name"] == prop].sort_values("total", ascending=False).head(5)
            cats_str = "; ".join(f"{r['Budget Category']}: ${r['total']:,.0f}" for _, r in p_cats.iterrows())
            lines.append(f"  {prop}: {cats_str}")

        # ── G. Unit Turnover Stats ──
        lines.append("\n=== REPEAT TURNERS — Units with Most Full Turns (Top 10) ===")
        unit_counts = _ft_turns.groupby(["Property Name", "Unit Label"]).agg(
            turns=("Turn Key", "count"),
            total_spend=("total_cost", "sum"),
            avg_cost=("total_cost", "mean"),
        ).sort_values("turns", ascending=False)
        for (prop, unit), row in unit_counts.head(10).iterrows():
            lines.append(f"  {prop} Unit {unit}: {row['turns']} turns, avg ${row['avg_cost']:,.0f}, total ${row['total_spend']:,.0f}")

        lines.append("\n=== HIGHEST-COST SINGLE TURNS (Top 10) ===")
        top_cost = _ft_turns.nlargest(10, "total_cost")
        for _, row in top_cost.iterrows():
            dt = row["Move-Out Date"].strftime("%b %Y") if pd.notna(row["Move-Out Date"]) else "—"
            lines.append(f"  {row['Property Name']} Unit {row['Unit Label']} ({dt}): ${row['total_cost']:,.0f}")

        # ── H. Floor Plan Detail by Property ──
        lines.append("\n=== FLOOR PLAN DETAIL BY PROPERTY (Full Turns) ===")
        fp_prop = _ft_turns.groupby(["Property Name", "Floor Plan"]).agg(
            turns=("Turn Key", "count"),
            avg_cost=("total_cost", "mean"),
        ).reset_index()
        for prop in sorted(_ft_turns["Property Name"].unique(), key=lambda n: _PROP_RANK.get(n, 999)):
            p_fp = fp_prop[fp_prop["Property Name"] == prop].sort_values("turns", ascending=False).head(3)
            fp_str = "; ".join(f"{r['Floor Plan']}: {r['turns']} turns, avg ${r['avg_cost']:,.0f}" for _, r in p_fp.iterrows())
            lines.append(f"  {prop}: {fp_str}")

        return "\n".join(lines)

    data_context = build_data_context(ft_turns, ft_lines, _df_all)

    SYSTEM_PROMPT = f"""You are a senior multifamily real estate analytics assistant embedded in a renovation dashboard for a 14-property multifamily real estate portfolio.
Your role is to answer questions about this portfolio's renovation data with precision, clarity, and executive-level insight.

Key definitions:
- Full Turn: A complete unit renovation after a tenant moves out — largest scope and cost
- Make Ready: A lighter-touch refresh between tenants — targeted scope (paint, cleaning, minor repairs)
- Partial Turn: Work done on a unit outside of a full move-out cycle
- Turn Key: Unique identifier for each turn event (Property + Unit + Move-Out Date)
- Duration: Days from move-out to last invoice (renovation timeline)
- Cost Types: Materials, Labor, Mixed, Fee — every invoice line is classified into one of these
- Budget Categories (17 total, grouped):
  Core Labor (7): Paint, Labor General, Flooring Labor, Electric General, Countertops Labor, Plumbing, Cabinets Labor
  Core Materials (4): Appliances, Flooring Materials, Cabinets Materials, Countertops Materials
  Other (6): Supplies, Powerwash and Demo, Management Fee, Scrape Ceiling, Glaze, Windows

Available data below includes:
- Portfolio overview and yearly trends
- Property breakdown and property × year detail (turn counts, avg cost, total spend)
- All 17 budget categories ranked by spend, plus category × year trends
- Cost type breakdown (Materials vs Labor vs Mixed vs Fee)
- Duration statistics by property
- All turn types (Full Turn, Make Ready, Partial Turn) — counts and costs
- Top categories by property, floor plan detail by property
- Repeat turners (units with most Full Turns) and highest-cost single turns
- Vendor summary and floor plan mix

Here is the portfolio data:

{data_context}

Guidelines:
- Always cite specific numbers from the data when answering
- Format currency as $X,XXX
- If you're unsure or the data doesn't support an answer, say so clearly
- Provide actionable insights when relevant
- Keep responses concise but thorough (3-5 sentences for simple questions, more for complex analysis)
- When comparing properties, always rank them by the relevant metric
- Use percentage changes for year-over-year comparisons
- Cross-reference dimensions when useful (e.g. property + year + category)
- If asked about a specific unit, note that you have aggregate data only — refer them to the Unit Search tab for unit-level detail"""

    # ── Provider & model config ──
    PROVIDERS = {
        "Claude (Anthropic)": {
            "models": ["claude-sonnet-4-20250514", "claude-haiku-4-5-20251001"],
            "default": "claude-sonnet-4-20250514",
            "placeholder": "sk-ant-...",
            "label": "Anthropic API Key",
        },
        "GPT (OpenAI)": {
            "models": ["gpt-4o-mini", "gpt-4o"],
            "default": "gpt-4o-mini",
            "placeholder": "sk-...",
            "label": "OpenAI API Key",
        },
        "Gemini (Google)": {
            "models": ["gemini-2.0-flash", "gemini-2.0-flash-lite"],
            "default": "gemini-2.0-flash",
            "placeholder": "AI...",
            "label": "Google AI API Key",
        },
    }

    col_prov, col_model = st.columns([1, 1])
    with col_prov:
        provider = st.selectbox("AI Provider", list(PROVIDERS.keys()))
    with col_model:
        prov_cfg = PROVIDERS[provider]
        model = st.selectbox("Model", prov_cfg["models"], index=0)

    api_key = st.text_input(
        prov_cfg["label"],
        type="password",
        placeholder=prov_cfg["placeholder"],
        help=f"Enter your {prov_cfg['label']} to enable AI Data Review. Your key is never stored.",
    )

    def call_llm(provider, model, api_key, system_prompt, messages):
        """Unified LLM call across providers. Returns the assistant response text."""
        if provider == "Claude (Anthropic)":
            client = Anthropic(api_key=api_key)
            # Anthropic uses a separate system param; messages must alternate user/assistant
            api_msgs = [{"role": m["role"], "content": m["content"]} for m in messages]
            response = client.messages.create(
                model=model,
                system=system_prompt,
                messages=api_msgs,
                max_tokens=1000,
                temperature=0.3,
            )
            return response.content[0].text

        elif provider == "GPT (OpenAI)":
            client = OpenAI(api_key=api_key)
            api_msgs = [{"role": "system", "content": system_prompt}] + [
                {"role": m["role"], "content": m["content"]} for m in messages
            ]
            response = client.chat.completions.create(
                model=model,
                messages=api_msgs,
                temperature=0.3,
                max_tokens=1000,
            )
            return response.choices[0].message.content

        elif provider == "Gemini (Google)":
            client = google_genai.Client(api_key=api_key)
            # Build contents: system instruction + conversation history
            contents = []
            for m in messages:
                role = "user" if m["role"] == "user" else "model"
                contents.append(google_genai.types.Content(
                    role=role,
                    parts=[google_genai.types.Part(text=m["content"])],
                ))
            response = client.models.generate_content(
                model=model,
                contents=contents,
                config=google_genai.types.GenerateContentConfig(
                    system_instruction=system_prompt,
                    temperature=0.3,
                    max_output_tokens=1000,
                ),
            )
            return response.text

    if not api_key:
        st.info(f"Enter your {prov_cfg['label']} above to start asking questions about your data. "
                "Your key is used only for this session and is never stored.")
        st.markdown("**Example questions you can ask:**")
        st.markdown("""
        - *Which property has the highest average Full Turn cost, and how has it trended?*
        - *Compare 2024 vs 2025 performance across the portfolio*
        - *What's our Materials vs Labor cost split?*
        - *Is paint getting more expensive year over year?*
        - *Which properties have the slowest turn durations?*
        - *What are the top cost drivers at Roscoe?*
        - *Which units are repeat turners — most Full Turns?*
        - *How many Make Readies vs Full Turns are we doing?*
        """)
        footer()
    else:
        # Clear chat when provider changes
        if "llm_provider" not in st.session_state:
            st.session_state.llm_provider = provider
        if st.session_state.llm_provider != provider:
            st.session_state.chat_messages = []
            st.session_state.llm_provider = provider

        # Initialize chat history
        if "chat_messages" not in st.session_state:
            st.session_state.chat_messages = []

        # Display chat history
        for msg in st.session_state.chat_messages:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        # Chat input
        if prompt := st.chat_input("Ask a question about your Full Turn data..."):
            # Show user message
            st.session_state.chat_messages.append({"role": "user", "content": prompt})
            with st.chat_message("user"):
                st.markdown(prompt)

            # Get AI response
            with st.chat_message("assistant"):
                with st.spinner(f"Analyzing via {provider}..."):
                    try:
                        # Include last 10 messages for context
                        recent = st.session_state.chat_messages[-10:]
                        answer = call_llm(provider, model, api_key, SYSTEM_PROMPT, recent)
                        st.markdown(answer)
                        st.session_state.chat_messages.append({"role": "assistant", "content": answer})
                    except Exception as e:
                        error_msg = str(e)
                        if "api_key" in error_msg.lower() or "auth" in error_msg.lower() or "api key" in error_msg.lower():
                            st.error(f"Invalid API key. Please check your {prov_cfg['label']} and try again.")
                        else:
                            st.error(f"Error: {error_msg}")

        # Clear chat button
        if st.session_state.chat_messages:
            if st.button("Clear Chat History"):
                st.session_state.chat_messages = []
                st.rerun()

        footer()
